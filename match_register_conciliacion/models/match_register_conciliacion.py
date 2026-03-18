# -*- coding: utf-8 -*-

from odoo import api, fields, models, _
from odoo.exceptions import UserError
from datetime import datetime, timedelta
import base64
import io
import logging
import re

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    openpyxl = None

try:
    import csv
except ImportError:
    csv = None

_logger = logging.getLogger(__name__)


class MatchRegisterConciliacion(models.Model):
    _name = 'match.register.conciliacion'
    _description = 'Registro de Conciliación con Matches'
    _order = 'date desc, name desc'

    name = fields.Char(
        string='Referencia',
        required=True,
        default=lambda self: _('Nuevo'),
        copy=False,
    )
    
    date = fields.Date(
        string='Fecha',
        required=True,
        default=fields.Date.today,
        index=True,
    )
    
    journal_id = fields.Many2one(
        'account.journal',
        string='Diario',
        required=True,
        domain=[('type', '=', 'bank')],
        index=True,
    )
    
    state = fields.Selection([
        ('draft', 'Borrador'),
        ('matches_found', 'Matches Encontrados'),
        ('reconciled', 'Reconciliado'),
    ], string='Estado', default='draft', required=True, index=True)
    
    # Líneas del match
    match_line_ids = fields.One2many(
        'match.register.conciliacion.line',
        'match_register_id',
        string='Líneas de Match',
    )
    
    match_count = fields.Integer(
        string='Cantidad de Matches',
        compute='_compute_match_count',
    )
    
    match_selected_count = fields.Integer(
        string='Matches Seleccionados',
        compute='_compute_match_count',
    )
    
    match_reconciled_count = fields.Integer(
        string='Matches Reconciliados',
        compute='_compute_match_count',
    )
    
    match_error_count = fields.Integer(
        string='Matches con Error',
        compute='_compute_match_count',
    )
    
    unreconcile_error_message = fields.Text(
        string='Errores al Romper Conciliación',
        readonly=True,
        help='Mensajes de error si hubo problemas al romper conciliaciones',
    )
    
    reconciled_move_ids = fields.Many2many(
        'account.move',
        string='Asientos Reconciliados',
        compute='_compute_reconciled_moves',
        help='Asientos contables que fueron reconciliados en este proceso',
    )
    
    company_id = fields.Many2one(
        'res.company',
        string='Compañía',
        related='journal_id.company_id',
        store=True,
        readonly=True,
    )
    
    currency_id = fields.Many2one(
        'res.currency',
        string='Moneda',
        related='journal_id.currency_id',
        store=True,
        readonly=True,
    )
    
    # Referencia al extracto bancario creado
    statement_id = fields.Many2one(
        'account.bank.statement',
        string='Extracto Bancario',
        readonly=True,
        help='Extracto bancario creado desde el archivo importado',
    )
    
    # Campos para importación de extracto
    import_file = fields.Binary(
        string='Archivo Excel/CSV',
        help='Suba el archivo Excel o CSV con el extracto bancario',
    )
    
    import_filename = fields.Char(string='Nombre del Archivo')
    
    import_file_type = fields.Selection([
        ('excel', 'Excel (.xlsx)'),
        ('csv', 'CSV'),
    ], string='Tipo de Archivo', default='excel')
    
    # Mapeo de columnas (formato simplificado como extracto bancario real)
    import_column_date = fields.Char(string='Columna Fecha', default='A', required=True)
    import_column_reference = fields.Char(string='Columna Referencia', default='B', required=True)
    import_column_amount = fields.Char(string='Columna Monto', default='D', required=True)
    import_column_description = fields.Char(string='Columna Descripción', default='C', 
                                            help='Opcional. Puede contener RIF/NIT o nombre del cliente')
    import_column_partner = fields.Char(string='Columna Cliente/Proveedor', default='', 
                                        help='Opcional. Si no está disponible, se buscará en la descripción')
    import_column_vat = fields.Char(string='Columna RIF/NIT', default='', 
                                   help='Opcional. Si no está disponible, se buscará en la descripción')
    import_start_row = fields.Integer(string='Fila de Inicio', default=2, 
                                     help='Fila donde comienzan los datos (1 = primera fila, normalmente 2 si hay encabezados)')
    import_date_format = fields.Char(string='Formato de Fecha', default='%d/%m/%Y',
                                    help='Formato de fecha en el Excel (ej: %d/%m/%Y, %Y-%m-%d)')
    
    # Método de match
    match_method = fields.Selection([
        ('exact_reference_only', 'Referencia Exacta (sin importar monto)'),
        ('partial_reference_only', 'Referencia Parcial (sin importar monto)'),
        ('exact_amount_reference', 'Monto Exacto + Referencia Exacta'),
        ('partial_reference_amount', 'Monto Exacto + Referencia Parcial'),
        ('vat_date_amount', 'VAT en Concepto + Fecha + Monto'),
        ('vat_amount', 'VAT en Concepto + Monto'),
        ('date_amount', 'Fecha + Monto'),
        ('partner_date_amount', 'Partner + Fecha + Monto'),
        ('partial_reference_date_amount', 'Parte de Referencia + Fecha + Monto'),
        ('vat_date_amount_tolerance', 'VAT en Concepto + Fecha + Monto (con tolerancia 5%)'),
        ('amount_tolerance', 'Monto con Tolerancia 5%'),
    ], string='Método de Match', default='exact_reference_only', required=True,
       help='Seleccione cómo desea buscar los matches entre el extracto y los pagos')
    
    @api.depends('match_line_ids', 'match_line_ids.selected', 'match_line_ids.state')
    def _compute_match_count(self):
        for record in self:
            record.match_count = len(record.match_line_ids)
            record.match_selected_count = len(record.match_line_ids.filtered(lambda m: m.selected))
            record.match_reconciled_count = len(record.match_line_ids.filtered(lambda m: m.state == 'reconciled'))
            record.match_error_count = len(record.match_line_ids.filtered(lambda m: m.state == 'error'))
    
    @api.depends('match_line_ids', 'match_line_ids.state', 'match_line_ids.payment_id')
    def _compute_reconciled_moves(self):
        """Calcular los asientos que fueron reconciliados"""
        for record in self:
            reconciled_matches = record.match_line_ids.filtered(lambda m: m.state == 'reconciled')
            moves = self.env['account.move']
            for match in reconciled_matches:
                if match.payment_id and match.payment_id.move_id:
                    moves |= match.payment_id.move_id
                    # También incluir los asientos de las líneas reconciliadas
                    for line in match.payment_id.move_id.line_ids:
                        if line.reconciled and line.full_reconcile_id:
                            # Obtener todos los asientos relacionados con esta reconciliación
                            for rec_line in line.full_reconcile_id.reconciled_line_ids:
                                if rec_line.move_id:
                                    moves |= rec_line.move_id
            record.reconciled_move_ids = moves
    
    @api.model
    def create(self, vals):
        if vals.get('name', _('Nuevo')) == _('Nuevo'):
            vals['name'] = self.env['ir.sequence'].next_by_code('match.register.conciliacion') or _('Nuevo')
        return super().create(vals)
    
    def action_generate_example_file(self):
        """Generar archivo Excel de ejemplo con datos ficticios"""
        self.ensure_one()
        
        if not self.journal_id:
            raise UserError(_('Debe seleccionar un diario primero.'))
        
        if not openpyxl:
            raise UserError(_('La librería openpyxl no está instalada. Por favor instálela con: pip install openpyxl'))
        
        # Crear workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Extracto Bancario"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados (formato simplificado como extracto bancario real)
        headers = ['Fecha', 'Referencia', 'Descripción', 'Monto']
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        # Datos de ejemplo ficticios (no datos reales)
        from datetime import datetime, timedelta
        today = fields.Date.today()
        
        example_data = [
            {
                'date': (today - timedelta(days=2)).strftime('%d/%m/%Y'),
                'reference': '52416549',
                'description': 'TRF CR INM - Cliente Ejemplo J-12345678-9',
                'amount': 1500.00,
            },
            {
                'date': (today - timedelta(days=1)).strftime('%d/%m/%Y'),
                'reference': '22374257',
                'description': 'PPV - Otro Cliente G-98765432-1',
                'amount': 200.00,
            },
            {
                'date': today.strftime('%d/%m/%Y'),
                'reference': '51822712116',
                'description': 'TRF.OB - Cliente Tercero V-11223344-5',
                'amount': 600.00,
            },
            {
                'date': today.strftime('%d/%m/%Y'),
                'reference': '13160595207',
                'description': 'Banesco Pago Movil - 18687408',
                'amount': 400.00,
            },
        ]
        
        # Escribir datos de ejemplo
        row = 2
        for data in example_data:
            sheet.cell(row=row, column=1, value=data['date'])
            sheet.cell(row=row, column=2, value=data['reference'])
            sheet.cell(row=row, column=3, value=data['description'])
            sheet.cell(row=row, column=4, value=data['amount'])
            row += 1
        
        # Ajustar ancho de columnas
        sheet.column_dimensions['A'].width = 12  # Fecha
        sheet.column_dimensions['B'].width = 20  # Referencia
        sheet.column_dimensions['C'].width = 50  # Descripción
        sheet.column_dimensions['D'].width = 15  # Monto
        
        # Guardar en memoria
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Codificar en base64
        file_data = base64.b64encode(output.read())
        filename = _('Extracto_Ejemplo_%s_%s.xlsx') % (
            self.journal_id.name.replace(' ', '_'),
            fields.Date.today().strftime('%Y%m%d')
        )
        
        # Crear attachment para descargar
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': file_data,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }
    
    def _column_to_index(self, column):
        """Convertir letra de columna a índice (A=0, B=1, etc.)"""
        if not column:
            return None
        column = column.upper().strip()
        index = 0
        for char in column:
            index = index * 26 + (ord(char) - ord('A') + 1)
        return index - 1
    
    def _parse_date(self, date_str, date_format, row_num=None):
        """Parsear fecha según el formato especificado"""
        if not date_str:
            if row_num:
                raise UserError(_('Fila %d: La fecha está vacía. Verifique que la columna de fecha esté correctamente mapeada.') % row_num)
            return None
        
        date_str_clean = str(date_str).strip()
        formats_to_try = [
            date_format,
            '%d/%m/%Y',
            '%Y-%m-%d',
            '%d-%m-%Y',
            '%Y/%m/%d',
            '%d.%m.%Y',
        ]
        
        for fmt in formats_to_try:
            try:
                return datetime.strptime(date_str_clean, fmt).date()
            except (ValueError, AttributeError):
                continue
        
        try:
            # Intentar parsear como fecha serial de Excel
            excel_epoch = datetime(1899, 12, 30)
            days = int(float(date_str))
            return (excel_epoch + timedelta(days=days)).date()
        except (ValueError, TypeError):
            pass
        
        # Mensaje de error detallado
        error_msg = _(
            'No se pudo parsear la fecha "%s" en la fila %d.\n\n'
            'Formato esperado: %s\n\n'
            'Formatos soportados:\n'
            '- %d/%m/%Y (ej: 17/11/2025)\n'
            '- %Y-%m-%d (ej: 2025-11-17)\n'
            '- %d-%m-%Y (ej: 17-11-2025)\n'
            '- %Y/%m/%d (ej: 2025/11/17)\n'
            '- %d.%m.%Y (ej: 17.11.2025)\n\n'
            'Verifique que:\n'
            '1. La columna de fecha esté correctamente mapeada\n'
            '2. El formato de fecha coincida con el formato en el Excel\n'
            '3. La fecha no esté vacía o con formato incorrecto'
        ) % (date_str_clean, row_num or '?', date_format)
        
        raise UserError(error_msg)
    
    def _create_bank_statement_from_lines(self, lines_data):
        """Crear extracto bancario desde las líneas procesadas"""
        self.ensure_one()
        
        if not self.journal_id:
            raise UserError(_('Debe seleccionar un diario antes de importar.'))
        
        # Crear extracto bancario
        statement = self.env['account.bank.statement'].create({
            'name': self.name or _('Extracto Importado - %s') % fields.Date.today(),
            'journal_id': self.journal_id.id,
            'date': self.date or fields.Date.today(),
        })
        
        # Crear líneas del extracto
        statement_lines = []
        for line_data in lines_data:
            line_vals = {
                'statement_id': statement.id,
                'date': line_data.get('date'),
                'payment_ref': line_data.get('reference', ''),
                'partner_id': line_data.get('partner_id', False),
                'amount': line_data.get('amount', 0.0),
                'narration': line_data.get('description', ''),
            }
            line = self.env['account.bank.statement.line'].create(line_vals)
            statement_lines.append(line)
        
        return statement, statement_lines
    
    def _extract_vat_from_description(self, description):
        """Extraer VAT/RIF/NIT de una descripción"""
        if not description:
            return None
        description_str = str(description)
        # Buscar formato J-12345678-9 o similar
        vat_pattern = re.search(r'[JGV]-?\s*(\d{6,9})-?\s*(\d)', description_str)
        if vat_pattern:
            return f"{vat_pattern.group(1)}-{vat_pattern.group(2)}"
        # Buscar solo números de 7-10 dígitos (posible RIF/NIT)
        vat_numbers = re.findall(r'\d{7,10}', description_str)
        if vat_numbers:
            return vat_numbers[0]
        return None
    
    def _find_matches_exact_reference_only(self, statement_line, payments, perfect_match_payment_ids):
        """Buscar matches por referencia exacta sin importar monto
        
        IMPORTANTE: Este método busca SOLO por referencia exacta.
        No requiere monto, fecha ni ningún otro criterio.
        """
        matching_payments = self.env['account.payment']
        is_perfect_match = False
        
        # Requerir que haya referencia en el extracto para hacer match
        if not statement_line.payment_ref:
            return matching_payments, is_perfect_match
        
        ref_extracto = statement_line.payment_ref.lower().strip()
        
        # Buscar matches donde la referencia sea exacta (sin importar monto)
        perfect_matches = payments.filtered(
            lambda p: p.id not in perfect_match_payment_ids and
                     ref_extracto == (p.memo or '').lower().strip()
        )
        
        if perfect_matches:
            matching_payments = perfect_matches
            is_perfect_match = True
        
        return matching_payments, is_perfect_match
    
    def _find_matches_partial_reference_only(self, statement_line, payments, perfect_match_payment_ids):
        """Buscar matches por referencia parcial sin importar monto
        
        IMPORTANTE: Este método busca SOLO por referencia parcial.
        No requiere monto, fecha ni ningún otro criterio.
        """
        matching_payments = self.env['account.payment']
        is_perfect_match = False
        
        if not statement_line.payment_ref:
            return matching_payments, is_perfect_match
        
        ref_extracto = statement_line.payment_ref.lower().strip()
        
        # Buscar pagos con referencia parcial (sin importar monto)
        for payment in payments:
            if payment.id in perfect_match_payment_ids:
                continue
            
            # Verificar referencia parcial
            payment_memo = (payment.memo or '').lower().strip()
            if ref_extracto in payment_memo or payment_memo in ref_extracto:
                matching_payments |= payment
                is_perfect_match = True
        
        return matching_payments, is_perfect_match
    
    def _find_matches_exact_amount_reference(self, statement_line, payments, perfect_match_payment_ids):
        """Buscar matches por monto exacto + referencia exacta
        
        IMPORTANTE: Este método requiere AMBOS criterios:
        - Monto exacto
        - Referencia exacta
        No hace fallback a solo monto, ya que el usuario seleccionó específicamente este método.
        """
        matching_payments = self.env['account.payment']
        is_perfect_match = False
        
        # Requerir que haya referencia en el extracto para hacer match
        if not statement_line.payment_ref:
            return matching_payments, is_perfect_match
        
        # Buscar matches donde AMBOS criterios se cumplan: monto exacto Y referencia exacta
        perfect_matches = payments.filtered(
            lambda p: p.id not in perfect_match_payment_ids and
                     abs(p.amount - abs(statement_line.amount)) < 0.01 and
                     statement_line.payment_ref.lower().strip() == (p.memo or '').lower().strip()
        )
        
        if perfect_matches:
            matching_payments = perfect_matches
            is_perfect_match = True
        
        # NO hacer fallback a solo monto - el usuario seleccionó específicamente "Monto Exacto + Referencia Exacta"
        return matching_payments, is_perfect_match
    
    def _find_matches_vat_date_amount(self, statement_line, payments, perfect_match_payment_ids):
        """Buscar matches por VAT en concepto + fecha + monto
        
        IMPORTANTE: Este método requiere TODOS los criterios:
        - VAT en la descripción del extracto
        - Fecha exacta
        - Monto exacto
        """
        matching_payments = self.env['account.payment']
        is_perfect_match = False
        
        # Extraer VAT de la descripción del extracto
        vat_extracto = self._extract_vat_from_description(statement_line.narration)
        if not vat_extracto:
            return matching_payments, is_perfect_match
        
        # Normalizar VAT
        vat_clean = vat_extracto.replace('-', '').replace('.', '').strip()
        
        # Buscar pagos con el mismo VAT, fecha y monto
        for payment in payments:
            if payment.id in perfect_match_payment_ids:
                continue
            
            # Verificar monto exacto
            if abs(payment.amount - abs(statement_line.amount)) >= 0.01:
                continue
            
            # Verificar fecha (misma fecha)
            if payment.date != statement_line.date:
                continue
            
            # Verificar VAT del partner del pago
            if payment.partner_id and payment.partner_id.vat:
                payment_vat = payment.partner_id.vat.replace('-', '').replace('.', '').strip()
                if vat_clean in payment_vat or payment_vat in vat_clean:
                    matching_payments |= payment
                    is_perfect_match = True
        
        return matching_payments, is_perfect_match
    
    def _find_matches_partial_reference_amount(self, statement_line, payments, perfect_match_payment_ids):
        """Buscar matches por parte de referencia + monto
        
        IMPORTANTE: Este método requiere AMBOS criterios:
        - Referencia parcial (la referencia del extracto debe estar contenida en el memo del pago o viceversa)
        - Monto exacto
        """
        matching_payments = self.env['account.payment']
        is_perfect_match = False
        
        if not statement_line.payment_ref:
            return matching_payments, is_perfect_match
        
        ref_extracto = statement_line.payment_ref.lower().strip()
        
        # Buscar pagos con monto exacto y referencia parcial
        for payment in payments:
            if payment.id in perfect_match_payment_ids:
                continue
            
            # Verificar monto exacto
            if abs(payment.amount - abs(statement_line.amount)) >= 0.01:
                continue
            
            # Verificar referencia parcial
            payment_memo = (payment.memo or '').lower().strip()
            if ref_extracto in payment_memo or payment_memo in ref_extracto:
                matching_payments |= payment
                is_perfect_match = True
        
        return matching_payments, is_perfect_match
    
    def _find_matches_vat_amount(self, statement_line, payments, perfect_match_payment_ids):
        """Buscar matches por VAT en concepto + monto
        
        IMPORTANTE: Este método requiere AMBOS criterios:
        - VAT en la descripción del extracto
        - Monto exacto
        """
        matching_payments = self.env['account.payment']
        is_perfect_match = False
        
        # Extraer VAT de la descripción del extracto
        vat_extracto = self._extract_vat_from_description(statement_line.narration)
        if not vat_extracto:
            return matching_payments, is_perfect_match
        
        # Normalizar VAT
        vat_clean = vat_extracto.replace('-', '').replace('.', '').strip()
        
        # Buscar pagos con el mismo VAT y monto
        for payment in payments:
            if payment.id in perfect_match_payment_ids:
                continue
            
            # Verificar monto exacto
            if abs(payment.amount - abs(statement_line.amount)) >= 0.01:
                continue
            
            # Verificar VAT del partner del pago
            if payment.partner_id and payment.partner_id.vat:
                payment_vat = payment.partner_id.vat.replace('-', '').replace('.', '').strip()
                if vat_clean in payment_vat or payment_vat in vat_clean:
                    matching_payments |= payment
                    is_perfect_match = True
        
        return matching_payments, is_perfect_match
    
    def _find_matches_date_amount(self, statement_line, payments, perfect_match_payment_ids):
        """Buscar matches por fecha + monto
        
        IMPORTANTE: Este método requiere AMBOS criterios:
        - Fecha exacta
        - Monto exacto
        """
        matching_payments = self.env['account.payment']
        is_perfect_match = False
        
        # Buscar pagos con la misma fecha y monto exacto
        matching_payments = payments.filtered(
            lambda p: p.id not in perfect_match_payment_ids and
                     p.date == statement_line.date and
                     abs(p.amount - abs(statement_line.amount)) < 0.01
        )
        
        if matching_payments:
            is_perfect_match = True
        
        return matching_payments, is_perfect_match
    
    def _find_matches_partner_date_amount(self, statement_line, payments, perfect_match_payment_ids):
        """Buscar matches por partner + fecha + monto
        
        IMPORTANTE: Este método requiere TODOS los criterios:
        - Partner exacto
        - Fecha exacta
        - Monto exacto
        """
        matching_payments = self.env['account.payment']
        is_perfect_match = False
        
        if not statement_line.partner_id:
            return matching_payments, is_perfect_match
        
        # Buscar pagos con el mismo partner, fecha y monto exacto
        matching_payments = payments.filtered(
            lambda p: p.id not in perfect_match_payment_ids and
                     p.partner_id == statement_line.partner_id and
                     p.date == statement_line.date and
                     abs(p.amount - abs(statement_line.amount)) < 0.01
        )
        
        if matching_payments:
            is_perfect_match = True
        
        return matching_payments, is_perfect_match
    
    def _find_matches_partial_reference_date_amount(self, statement_line, payments, perfect_match_payment_ids):
        """Buscar matches por parte de referencia + fecha + monto
        
        IMPORTANTE: Este método requiere TODOS los criterios:
        - Referencia parcial (la referencia del extracto debe estar contenida en el memo del pago o viceversa)
        - Fecha exacta
        - Monto exacto
        """
        matching_payments = self.env['account.payment']
        is_perfect_match = False
        
        if not statement_line.payment_ref:
            return matching_payments, is_perfect_match
        
        ref_extracto = statement_line.payment_ref.lower().strip()
        
        # Buscar pagos con monto exacto, fecha y referencia parcial
        for payment in payments:
            if payment.id in perfect_match_payment_ids:
                continue
            
            # Verificar monto exacto
            if abs(payment.amount - abs(statement_line.amount)) >= 0.01:
                continue
            
            # Verificar fecha
            if payment.date != statement_line.date:
                continue
            
            # Verificar referencia parcial
            payment_memo = (payment.memo or '').lower().strip()
            if ref_extracto in payment_memo or payment_memo in ref_extracto:
                matching_payments |= payment
                is_perfect_match = True
        
        return matching_payments, is_perfect_match
    
    def _find_matches_vat_date_amount_tolerance(self, statement_line, payments, perfect_match_payment_ids):
        """Buscar matches por VAT en concepto + fecha + monto con tolerancia 5%
        
        IMPORTANTE: Este método requiere TODOS los criterios:
        - VAT en la descripción del extracto
        - Fecha exacta
        - Monto con tolerancia del 5%
        """
        matching_payments = self.env['account.payment']
        is_perfect_match = False
        
        # Extraer VAT de la descripción del extracto
        vat_extracto = self._extract_vat_from_description(statement_line.narration)
        if not vat_extracto:
            return matching_payments, is_perfect_match
        
        # Normalizar VAT
        vat_clean = vat_extracto.replace('-', '').replace('.', '').strip()
        
        # Buscar pagos con el mismo VAT, fecha y monto (con tolerancia 5%)
        for payment in payments:
            if payment.id in perfect_match_payment_ids:
                continue
            
            # Verificar monto con tolerancia 5%
            amount_diff = abs(payment.amount - abs(statement_line.amount)) / max(abs(statement_line.amount), 1)
            if amount_diff >= 0.05:
                continue
            
            # Verificar fecha (misma fecha)
            if payment.date != statement_line.date:
                continue
            
            # Verificar VAT del partner del pago
            if payment.partner_id and payment.partner_id.vat:
                payment_vat = payment.partner_id.vat.replace('-', '').replace('.', '').strip()
                if vat_clean in payment_vat or payment_vat in vat_clean:
                    matching_payments |= payment
                    is_perfect_match = True
        
        return matching_payments, is_perfect_match
    
    def _find_matches_amount_tolerance(self, statement_line, payments, perfect_match_payment_ids):
        """Buscar matches por monto con tolerancia 5%
        
        IMPORTANTE: Este método busca solo por monto con tolerancia del 5%.
        No requiere otros criterios adicionales.
        """
        matching_payments = self.env['account.payment']
        is_perfect_match = False
        
        # Buscar pagos con monto aproximado (5% tolerancia)
        matching_payments = payments.filtered(
            lambda p: p.id not in perfect_match_payment_ids and
                     abs(p.amount - abs(statement_line.amount)) / max(abs(statement_line.amount), 1) < 0.05
        )
        
        if matching_payments:
            is_perfect_match = True
        
        return matching_payments, is_perfect_match
    
    def _find_matches_by_method(self, statement_line, payments, perfect_match_payment_ids):
        """Buscar matches según el método seleccionado"""
        method = self.match_method
        
        if method == 'exact_reference_only':
            return self._find_matches_exact_reference_only(statement_line, payments, perfect_match_payment_ids)
        elif method == 'partial_reference_only':
            return self._find_matches_partial_reference_only(statement_line, payments, perfect_match_payment_ids)
        elif method == 'exact_amount_reference':
            return self._find_matches_exact_amount_reference(statement_line, payments, perfect_match_payment_ids)
        elif method == 'partial_reference_amount':
            return self._find_matches_partial_reference_amount(statement_line, payments, perfect_match_payment_ids)
        elif method == 'vat_date_amount':
            return self._find_matches_vat_date_amount(statement_line, payments, perfect_match_payment_ids)
        elif method == 'vat_amount':
            return self._find_matches_vat_amount(statement_line, payments, perfect_match_payment_ids)
        elif method == 'date_amount':
            return self._find_matches_date_amount(statement_line, payments, perfect_match_payment_ids)
        elif method == 'partner_date_amount':
            return self._find_matches_partner_date_amount(statement_line, payments, perfect_match_payment_ids)
        elif method == 'partial_reference_date_amount':
            return self._find_matches_partial_reference_date_amount(statement_line, payments, perfect_match_payment_ids)
        elif method == 'vat_date_amount_tolerance':
            return self._find_matches_vat_date_amount_tolerance(statement_line, payments, perfect_match_payment_ids)
        elif method == 'amount_tolerance':
            return self._find_matches_amount_tolerance(statement_line, payments, perfect_match_payment_ids)
        else:
            # Por defecto, usar exact_reference_only
            return self._find_matches_exact_reference_only(statement_line, payments, perfect_match_payment_ids)
    
    def action_import_and_find_matches(self):
        """Importar extracto y buscar matches automáticamente"""
        self.ensure_one()
        
        if not self.import_file:
            raise UserError(_('Por favor suba un archivo Excel o CSV.'))
        
        if self.state != 'draft':
            raise UserError(_('Solo se puede importar en estado Borrador.'))
        
        if not self.journal_id:
            raise UserError(_('Debe seleccionar un diario antes de importar.'))
        
        # Limpiar matches anteriores
        self.match_line_ids.unlink()
        
        # Procesar archivo
        if self.import_file_type == 'excel':
            if not openpyxl:
                raise UserError(_('La librería openpyxl no está instalada. Por favor instálela con: pip install openpyxl'))
            lines_data = self._process_excel()
        else:
            if not csv:
                raise UserError(_('El módulo csv no está disponible.'))
            lines_data = self._process_csv()
        
        # Crear extracto bancario desde las líneas procesadas
        statement, statement_lines = self._create_bank_statement_from_lines(lines_data)
        
        # Guardar referencia al extracto bancario
        self.statement_id = statement.id
        
        # Buscar matches directamente (sin depender de inherit_account_payment_register)
        matches_created = []
        perfect_match_payment_ids = set()
        
        for statement_line in statement_lines:
            if not statement_line.amount:
                continue
            
            # Buscar pagos del mismo diario en proceso
            domain = [
                ('journal_id', '=', self.journal_id.id),
                ('state', '=', 'in_process'),
            ]
            
            # Excluir pagos que ya tienen un match perfecto
            if perfect_match_payment_ids:
                domain.append(('id', 'not in', list(perfect_match_payment_ids)))
            
            payments = self.env['account.payment'].search(domain)
            
            # Buscar matches según el método seleccionado
            matching_payments, is_perfect_match = self._find_matches_by_method(
                statement_line, payments, perfect_match_payment_ids
            )
            
            # Crear matches en nuestro modelo
            for payment in matching_payments:
                match_vals = {
                    'match_register_id': self.id,
                    'payment_id': payment.id,
                    'extracto_date': statement_line.date,
                    'extracto_reference': statement_line.payment_ref or '',
                    'extracto_description': statement_line.narration or '',
                    'extracto_amount': statement_line.amount,
                    'extracto_partner_id': statement_line.partner_id.id if statement_line.partner_id else False,
                    'selected': True,
                }
                match = self.env['match.register.conciliacion.line'].create(match_vals)
                matches_created.append(match)
                
                # Si es match perfecto, marcar el pago
                if is_perfect_match:
                    perfect_match_payment_ids.add(payment.id)
        
        # Actualizar estado
        if matches_created:
            self.state = 'matches_found'
            # No mostrar notificación, solo actualizar el estado
            return {'type': 'ir.actions.act_window_close'}
        else:
            raise UserError(_('Se importaron %d líneas pero no se encontraron matches.') % len(lines_data))
    
    def _process_excel(self):
        """Procesar archivo Excel"""
        try:
            file_data = base64.b64decode(self.import_file)
            workbook = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError(_('Error al leer el archivo Excel: %s') % str(e))
        
        col_date_idx = self._column_to_index(self.import_column_date)
        col_ref_idx = self._column_to_index(self.import_column_reference)
        col_partner_idx = self._column_to_index(self.import_column_partner) if self.import_column_partner else None
        col_amount_idx = self._column_to_index(self.import_column_amount)
        col_desc_idx = self._column_to_index(self.import_column_description) if self.import_column_description else None
        col_vat_idx = self._column_to_index(self.import_column_vat) if self.import_column_vat else None
        
        if col_date_idx is None or col_ref_idx is None or col_amount_idx is None:
            raise UserError(_('Debe definir al menos las columnas: Fecha, Referencia y Monto.'))
        
        lines_data = []
        start_row = max(1, self.import_start_row)
        rows_processed = 0
        rows_skipped = 0
        rows_with_errors = 0
        
        for row_idx in range(start_row, sheet.max_row + 1):
            try:
                date_val = sheet.cell(row=row_idx, column=col_date_idx + 1).value
                ref_val = sheet.cell(row=row_idx, column=col_ref_idx + 1).value
                partner_val = sheet.cell(row=row_idx, column=col_partner_idx + 1).value if col_partner_idx is not None else None
                amount_val = sheet.cell(row=row_idx, column=col_amount_idx + 1).value
                desc_val = sheet.cell(row=row_idx, column=col_desc_idx + 1).value if col_desc_idx is not None else None
                vat_val = sheet.cell(row=row_idx, column=col_vat_idx + 1).value if col_vat_idx is not None else None
                
                # Validación mínima: fecha, referencia y monto
                if not date_val or not ref_val or amount_val is None:
                    _logger.warning("Fila %d: Saltando - datos incompletos (fecha=%s, ref=%s, monto=%s)", 
                                row_idx, date_val, ref_val, amount_val)
                    rows_skipped += 1
                    continue
                
                try:
                    date = self._parse_date(date_val, self.import_date_format, row_num=row_idx)
                except UserError:
                    # Re-lanzar UserError para que se muestre al usuario
                    raise
                except Exception as e:
                    error_msg = _('Fila %d: Error inesperado al parsear la fecha "%s": %s') % (row_idx, date_val, str(e))
                    _logger.warning("Fila %d: Error parseando fecha '%s' con formato '%s': %s", 
                                row_idx, date_val, self.import_date_format, str(e))
                    rows_with_errors += 1
                    # Continuar con la siguiente fila en lugar de fallar todo
                    continue
                
                try:
                    # Manejar montos con formato de texto (ej: "1.600,00" o "-200,00")
                    amount_str = str(amount_val).strip().replace('.', '').replace(',', '.')
                    amount = float(amount_str)
                except (ValueError, TypeError) as e:
                    _logger.warning("Fila %d: Error parseando monto '%s': %s", row_idx, amount_val, str(e))
                    rows_with_errors += 1
                    continue
                
                if amount == 0:
                    rows_skipped += 1
                    continue
                
                # Buscar partner: primero en columna dedicada, luego en descripción (RIF/NIT), luego en referencia
                partner_id = False
                partner = False
                
                # 1. Si hay columna de partner, buscar ahí
                if partner_val:
                    partner = self.env['res.partner'].search([
                        ('name', 'ilike', str(partner_val).strip())
                    ], limit=1)
                
                # 2. Si hay columna de VAT o descripción contiene RIF/NIT, buscar por VAT
                if not partner:
                    vat_to_search = None
                    if vat_val:
                        vat_to_search = str(vat_val).strip()
                    elif desc_val:
                        # Buscar RIF/NIT en la descripción (formato común: J-12345678-9 o 12345678)
                        vat_pattern = re.search(r'[JGV]-?\s*(\d{6,9})-?\s*(\d)', str(desc_val))
                        if vat_pattern:
                            vat_to_search = f"{vat_pattern.group(1)}-{vat_pattern.group(2)}"
                        else:
                            # Buscar solo números de 7-10 dígitos (posible RIF/NIT)
                            vat_numbers = re.findall(r'\d{7,10}', str(desc_val))
                            if vat_numbers:
                                vat_to_search = vat_numbers[0]
                    
                    if vat_to_search:
                        # Normalizar formato de VAT
                        vat_clean = vat_to_search.replace('-', '').replace('.', '').strip()
                        if len(vat_clean) >= 7:
                            # Buscar con diferentes formatos
                            partner = self.env['res.partner'].search([
                                '|',
                                ('vat', '=', vat_clean),
                                ('vat', 'ilike', vat_clean),
                            ], limit=1)
                
                # 3. Si no se encontró partner, intentar buscar por referencia (puede contener número de factura)
                if not partner and ref_val:
                    # Buscar partner que tenga esta referencia en sus facturas
                    # En account.move (facturas), el campo es payment_reference, no memo
                    invoice = self.env['account.move'].search([
                        ('payment_reference', 'ilike', str(ref_val).strip()),
                        ('partner_id', '!=', False),
                    ], limit=1, order='date desc')
                    if invoice:
                        partner = invoice.partner_id
                
                if partner:
                    partner_id = partner.id
                
                lines_data.append({
                    'date': date,
                    'reference': str(ref_val).strip() if ref_val else '',
                    'partner_id': partner_id,
                    'amount': amount,
                    'description': str(desc_val).strip() if desc_val else (str(partner_val).strip() if partner_val else ''),
                })
                rows_processed += 1
            except Exception as e:
                rows_with_errors += 1
                _logger.error("Fila %d: ✗ Error procesando: %s", row_idx, str(e), exc_info=True)
        
        _logger.info("Resumen: %d líneas procesadas, %d saltadas, %d con errores, Total líneas válidas: %d", 
                    rows_processed, rows_skipped, rows_with_errors, len(lines_data))
        
        if len(lines_data) == 0:
            error_details = []
            error_details.append(_('No se procesó ninguna línea válida del archivo Excel.'))
            error_details.append('')
            error_details.append(_('Configuración actual:'))
            error_details.append(_('  - Fila de inicio: %d') % start_row)
            error_details.append(_('  - Total de filas en hoja: %d') % sheet.max_row)
            error_details.append(_('  - Mapeo de columnas: Fecha=%s, Referencia=%s, Monto=%s') % (
                self.import_column_date, self.import_column_reference, self.import_column_amount))
            error_details.append('')
            
            if rows_with_errors > 0:
                error_details.append(_('Se encontraron %d filas con errores.') % rows_with_errors)
                error_details.append(_('Revisa los logs del servidor para ver los detalles de cada error.'))
                error_details.append('')
                error_details.append(_('Errores comunes:'))
                error_details.append(_('  - Formato de fecha incorrecto (verificar formato configurado)'))
                error_details.append(_('  - Monto no numérico o con formato incorrecto'))
                error_details.append(_('  - Fechas, referencias o montos vacíos'))
            elif rows_skipped > 0:
                error_details.append(_('Se saltaron %d filas porque tenían datos incompletos (fecha, referencia o monto vacíos).') % rows_skipped)
            else:
                error_details.append(_('No se encontraron datos en las filas especificadas.'))
                error_details.append(_('Verifica que:'))
                error_details.append(_('  - La fila de inicio sea correcta (normalmente 2 si hay encabezados)'))
                error_details.append(_('  - Las columnas estén correctamente mapeadas'))
                error_details.append(_('  - El archivo tenga datos en las filas esperadas'))
            
            error_msg = '\n'.join(error_details)
            _logger.warning("⚠️ %s", error_msg)
            raise UserError(error_msg)
        
        return lines_data
    
    def _process_csv(self):
        """Procesar archivo CSV"""
        try:
            file_data = base64.b64decode(self.import_file)
            content = file_data.decode('utf-8')
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)
        except Exception as e:
            raise UserError(_('Error al leer el archivo CSV: %s') % str(e))
        
        try:
            col_date_idx = int(self.import_column_date) - 1 if self.import_column_date.isdigit() else self._column_to_index(self.import_column_date)
            col_ref_idx = int(self.import_column_reference) - 1 if self.import_column_reference.isdigit() else self._column_to_index(self.import_column_reference)
            col_partner_idx = int(self.import_column_partner) - 1 if self.import_column_partner.isdigit() else self._column_to_index(self.import_column_partner)
            col_amount_idx = int(self.import_column_amount) - 1 if self.import_column_amount.isdigit() else self._column_to_index(self.import_column_amount)
            col_desc_idx = int(self.import_column_description) - 1 if self.import_column_description.isdigit() else self._column_to_index(self.import_column_description)
            col_vat_idx = int(self.import_column_vat) - 1 if self.import_column_vat and self.import_column_vat.isdigit() else (self._column_to_index(self.import_column_vat) if self.import_column_vat else None)
        except (ValueError, AttributeError):
            raise UserError(_('Las columnas deben ser números (1, 2, 3...) para archivos CSV.'))
        
        if col_date_idx is None or col_ref_idx is None or col_partner_idx is None or col_amount_idx is None:
            raise UserError(_('Debe definir al menos las columnas: Fecha, Referencia, Cliente/Proveedor y Monto.'))
        
        lines_data = []
        start_row = max(0, self.import_start_row - 1)
        
        for row_idx, row in enumerate(rows[start_row:], start=start_row + 1):
            try:
                if len(row) <= max(col_date_idx, col_ref_idx, col_partner_idx, col_amount_idx):
                    continue
                
                date_val = row[col_date_idx] if col_date_idx < len(row) else None
                ref_val = row[col_ref_idx] if col_ref_idx < len(row) else None
                partner_val = row[col_partner_idx] if col_partner_idx < len(row) else None
                amount_val = row[col_amount_idx] if col_amount_idx < len(row) else None
                desc_val = row[col_desc_idx] if col_desc_idx is not None and col_desc_idx < len(row) else ''
                vat_val = row[col_vat_idx] if col_vat_idx is not None and col_vat_idx < len(row) else ''
                
                if not date_val or not ref_val or not amount_val:
                    continue
                
                try:
                    date = self._parse_date(date_val, self.import_date_format)
                except Exception:
                    continue
                
                try:
                    amount = float(amount_val)
                except (ValueError, TypeError):
                    continue
                
                if amount == 0:
                    continue
                
                partner_id = False
                if partner_val:
                    partner = self.env['res.partner'].search([
                        '|', '|',
                        ('name', 'ilike', str(partner_val).strip()),
                        ('vat', '=', str(vat_val).strip()) if vat_val else (1, '=', 0),
                        ('ref', '=', str(ref_val).strip()),
                    ], limit=1)
                    
                    if not partner and vat_val:
                        partner = self.env['res.partner'].search([
                            ('vat', '=', str(vat_val).strip())
                        ], limit=1)
                    
                    if partner:
                        partner_id = partner.id
                
                lines_data.append({
                    'date': date,
                    'reference': str(ref_val).strip() if ref_val else '',
                    'partner_id': partner_id,
                    'amount': amount,
                    'description': str(desc_val).strip() if desc_val else '',
                })
                
            except Exception as e:
                _logger.error("Error procesando fila %d: %s", row_idx, str(e), exc_info=True)
        
        return lines_data
    
    def action_reset_to_draft(self):
        """Volver a estado Borrador para buscar nuevos matches"""
        self.ensure_one()
        
        # Limpiar matches anteriores pero mantener la configuración de importación
        self.match_line_ids.unlink()
        self.state = 'draft'
        
        # No mostrar notificación, solo actualizar el estado
        return {'type': 'ir.actions.act_window_close'}
    
    def action_find_matches(self):
        """Buscar matches con pagos pendientes"""
        self.ensure_one()
        
        if self.state != 'draft':
            raise UserError(_('Solo se pueden buscar matches en estado Borrador.'))
        
        # Limpiar matches anteriores
        self.match_line_ids.unlink()
        
        # Buscar pagos del diario en proceso de pago (simple y directo)
        payments = self.env['account.payment'].search([
            ('journal_id', '=', self.journal_id.id),
            ('state', '=', 'in_process'),
        ])
        
        _logger.info("Buscando matches para diario %s (ID: %s). Pagos encontrados: %d", 
                    self.journal_id.name, self.journal_id.id, len(payments))
        
        matches_created = []
        for payment in payments:
            match_vals = {
                'match_register_id': self.id,
                'payment_id': payment.id,
                'amount': payment.amount,
                'partner_id': payment.partner_id.id if payment.partner_id else False,
                'payment_memo': payment.memo or '',
                'date': payment.date,
                'selected': True,  # Por defecto seleccionados
            }
            match = self.env['match.register.conciliacion.line'].create(match_vals)
            matches_created.append(match)
        
        if matches_created:
            self.state = 'matches_found'
            # No mostrar notificación, solo actualizar el estado
            return {'type': 'ir.actions.act_window_close'}
        else:
            raise UserError(_('No se encontraron pagos pendientes para este diario.'))
    
    def action_confirm_and_reconcile(self):
        """Confirmar matches y reconciliar directamente sin crear extracto"""
        self.ensure_one()
        
        if self.state != 'matches_found':
            raise UserError(_('Solo se pueden confirmar matches en estado "Matches Encontrados".'))
        
        selected_matches = self.match_line_ids.filtered(lambda m: m.selected)
        if not selected_matches:
            raise UserError(_('Debe seleccionar al menos un match para confirmar.'))
        
        # Validar que no haya múltiples matches seleccionados para el mismo pago
        # Agrupar matches por pago
        matches_by_payment = {}
        for match in selected_matches:
            payment_id = match.payment_id.id
            if payment_id not in matches_by_payment:
                matches_by_payment[payment_id] = []
            matches_by_payment[payment_id].append(match)
        
        # Validar cada pago
        for payment_id, payment_matches in matches_by_payment.items():
            payment = payment_matches[0].payment_id
            payment_amount = payment.amount
            
            # Si hay más de 1 match seleccionado para el mismo pago, validar
            if len(payment_matches) > 1:
                # Obtener todos los extractos para este pago (en este registro)
                extracto_amounts = [m.extracto_amount for m in payment_matches]
                total_extracto = sum(extracto_amounts)
                
                # Verificar si este pago está en otros registros reconciliados
                other_registers = self.env['match.register.conciliacion'].search([
                    ('id', '!=', self.id),
                    ('match_line_ids.payment_id', '=', payment_id),
                    ('match_line_ids.state', '=', 'reconciled'),
                ])
                
                if other_registers:
                    # Si está en otros registros, sumar todos los extractos de todos los registros
                    for other_register in other_registers:
                        other_matches = other_register.match_line_ids.filtered(
                            lambda m: m.payment_id.id == payment_id and m.state == 'reconciled'
                        )
                        for other_match in other_matches:
                            total_extracto += other_match.extracto_amount
                
                # VALIDACIÓN ESTRICTA: El total de extractos DEBE ser EXACTAMENTE igual al monto del pago
                # El pago puede reconciliarse con múltiples facturas (OCA lo maneja automáticamente)
                # pero el extracto debe coincidir exactamente con el monto del pago
                if abs(total_extracto - payment_amount) > 0.01:
                    # Construir mensaje detallado con información de cada match
                    matches_info = []
                    for m in payment_matches:
                        matches_info.append(
                            _('  - Extracto: Ref=%s, Desc=%s, Monto=%s') % (
                                m.extracto_reference or 'N/A',
                                m.extracto_description or 'N/A',
                                m.extracto_amount
                            )
                        )
                    
                    diferencia = total_extracto - payment_amount
                    if diferencia > 0.01:
                        mensaje_error = _(
                            '⚠️ ALERTA: El total de los extractos (%s) EXCEDE el monto del pago %s (%s).\n\n'
                            'Diferencia: %s\n\n'
                            '❌ NO SE PUEDE RECONCILIAR: Los montos deben coincidir EXACTAMENTE.\n\n'
                            'Nota: El pago puede reconciliarse con múltiples facturas, pero el total del extracto '
                            'debe ser igual al monto del pago.\n\n'
                            'Detalle de extractos seleccionados:\n%s\n\n'
                            'Por favor, ajuste la selección para que el total de extractos sea igual al monto del pago.'
                        ) % (
                            total_extracto, payment.name, payment_amount, diferencia,
                            '\n'.join(matches_info)
                        )
                    else:
                        mensaje_error = _(
                            '⚠️ ALERTA: El total de los extractos (%s) es MENOR que el monto del pago %s (%s).\n\n'
                            'Diferencia: %s\n\n'
                            '❌ NO SE PUEDE RECONCILIAR: Los montos deben coincidir EXACTAMENTE.\n\n'
                            'Nota: El pago puede reconciliarse con múltiples facturas, pero el total del extracto '
                            'debe ser igual al monto del pago.\n\n'
                            'Detalle de extractos seleccionados:\n%s\n\n'
                            'Por favor, seleccione extractos adicionales o ajuste la selección.'
                        ) % (
                            total_extracto, payment.name, payment_amount, abs(diferencia),
                            '\n'.join(matches_info)
                        )
                    
                    raise UserError(mensaje_error)
                
                # Validar que no haya extractos duplicados problemáticos
                # PERMITIR múltiples extractos si suman exactamente el monto del pago
                # Ejemplo válido: pago de 5000 puede reconciliarse con 5 extractos de 1000 cada uno
                # Ejemplo inválido: pago de 200 con 2 extractos de 200 (duplicados)
                amount_counts = {}
                for amount in extracto_amounts:
                    amount_counts[amount] = amount_counts.get(amount, 0) + 1
                
                # Solo validar duplicados si un extracto individual coincide exactamente con el pago
                # y hay múltiples extractos del mismo monto (esto indicaría duplicados)
                for amount, count in amount_counts.items():
                    if count > 1 and abs(amount - payment_amount) < 0.01:
                        # Hay múltiples extractos del mismo monto que coincide con el pago
                        # Esto es sospechoso (duplicados), pero permitirlo si la suma total es correcta
                        # (ya validamos arriba que total_extracto == payment_amount)
                        _logger.info(
                            "Múltiples extractos del mismo monto (%s) para pago %s, pero la suma total es correcta. "
                            "Permitiendo reconciliación.",
                            amount, payment.name
                        )
                
                # Esta validación ya se hace arriba, pero la mantenemos como doble verificación
                # NO se permiten extractos mayores que el pago
        
        # VALIDACIÓN CRÍTICA: Agrupar matches por extracto y validar que la suma de pagos coincida
        # Cada línea del Excel es un extracto independiente
        # Si un extracto tiene múltiples pagos, la suma de los pagos debe ser igual al extracto
        matches_by_extracto = {}
        for match in selected_matches:
            extracto_key = (match.extracto_reference or '', match.extracto_amount)
            if extracto_key not in matches_by_extracto:
                matches_by_extracto[extracto_key] = []
            matches_by_extracto[extracto_key].append(match)
        
        # Validar cada extracto
        for extracto_key, extracto_matches in matches_by_extracto.items():
            extracto_ref, extracto_amount = extracto_key
            
            # Si hay más de 1 match para el mismo extracto, validar que la suma de pagos coincida
            if len(extracto_matches) > 1:
                payment_amounts = [abs(m.payment_id.amount) for m in extracto_matches]
                total_pagos = sum(payment_amounts)
                
                if abs(total_pagos - abs(extracto_amount)) > 0.01:
                    # Construir mensaje detallado con información de cada pago
                    pagos_info = []
                    for m in extracto_matches:
                        pagos_info.append(
                            _('  - Pago: %s, Monto=%s, Memo=%s') % (
                                m.payment_id.name,
                                m.payment_id.amount,
                                m.payment_memo or 'Sin memo'
                            )
                        )
                    
                    diferencia = total_pagos - abs(extracto_amount)
                    if diferencia > 0.01:
                        mensaje_error = _(
                            '❌ ERROR: El total de los pagos (%s) EXCEDE el monto del extracto (%s).\n\n'
                            'Extracto: Ref=%s, Monto=%s\n'
                            'Total de pagos seleccionados: %s\n'
                            'Diferencia: %s\n\n'
                            '⚠️ VERIFIQUE LOS PAGOS: Hay pagos que al parecer tienen mal el monto o la referencia '
                            'para poder cuadrar el asiento de conciliación.\n\n'
                            'Detalle de pagos seleccionados:\n%s\n\n'
                            'Por favor, verifique que los montos y referencias de los pagos sean correctos.'
                        ) % (
                            total_pagos, abs(extracto_amount),
                            extracto_ref or 'N/A', extracto_amount,
                            total_pagos, diferencia,
                            '\n'.join(pagos_info)
                        )
                    else:
                        mensaje_error = _(
                            '❌ ERROR: El total de los pagos (%s) es MENOR que el monto del extracto (%s).\n\n'
                            'Extracto: Ref=%s, Monto=%s\n'
                            'Total de pagos seleccionados: %s\n'
                            'Diferencia: %s\n\n'
                            '⚠️ VERIFIQUE LOS PAGOS: Hay pagos que al parecer tienen mal el monto o la referencia '
                            'para poder cuadrar el asiento de conciliación.\n\n'
                            'Detalle de pagos seleccionados:\n%s\n\n'
                            'Por favor, verifique que los montos y referencias de los pagos sean correctos.'
                        ) % (
                            total_pagos, abs(extracto_amount),
                            extracto_ref or 'N/A', extracto_amount,
                            total_pagos, abs(diferencia),
                            '\n'.join(pagos_info)
                        )
                    
                    raise UserError(mensaje_error)
        
        # Crear un conjunto de extractos que tienen múltiples pagos (para saltar validación individual)
        extractos_con_multiples_pagos = set()
        for extracto_key, extracto_matches in matches_by_extracto.items():
            if len(extracto_matches) > 1:
                extractos_con_multiples_pagos.add(extracto_key)
        
        # Reconciliar usando el extracto bancario creado
        # El extracto bancario tiene las líneas que se reconcilian con los pagos
        reconciled_count = [0]  # Usar lista para poder modificar desde métodos auxiliares
        errors = []
        
        # Reconciliar usando el extracto bancario (como inherit_account_payment_register)
        # Las líneas del extracto se reconcilian con las líneas de los pagos
        if not self.statement_id:
            raise UserError(_('No se encontró el extracto bancario. Por favor, importe el extracto nuevamente.'))
        
        statement = self.statement_id
        
        # Agrupar matches por extracto para reconciliar juntos cuando hay múltiples pagos
        # Esto permite reconciliar un extracto de 300 con 3 pagos de 50, 50, 200
        matches_por_extracto = {}
        for match in selected_matches:
            extracto_key = (match.extracto_reference or '', match.extracto_amount)
            if extracto_key not in matches_por_extracto:
                matches_por_extracto[extracto_key] = []
            matches_por_extracto[extracto_key].append(match)
        
        # Reconciliar agrupando por extracto
        for extracto_key, extracto_matches in matches_por_extracto.items():
            # Si hay múltiples pagos para este extracto, reconciliarlos todos juntos
            if len(extracto_matches) > 1:
                # Reconciliar todos los pagos de este extracto juntos
                self._reconcile_multiples_pagos_extracto(extracto_matches, statement, extractos_con_multiples_pagos, errors, reconciled_count)
            else:
                # Un solo pago, reconciliar normalmente
                match = extracto_matches[0]
                try:
                    self._reconcile_single_match(match, statement, extractos_con_multiples_pagos, errors, reconciled_count)
                except Exception as e:
                    error_msg = _('Error en match %s: %s') % (match.payment_id.name, str(e))
                    match.state = 'error'
                    errors.append(error_msg)
                    _logger.error("✗ ERROR EXCEPCIÓN en match %s: %s", match.id, str(e), exc_info=True)
        
        # Actualizar estado
        if reconciled_count[0] == len(selected_matches):
            self.state = 'reconciled'
        elif reconciled_count[0] > 0:
            self.state = 'reconciled'  # Si al menos uno se reconcilió, marcamos como reconciliado
        
        return {'type': 'ir.actions.act_window_close'}
    
    def _reconcile_single_match(self, match, statement, extractos_con_multiples_pagos, errors, reconciled_count):
        """Reconciliar un match individual (un pago con un extracto)"""
        payment = match.payment_id
        
        # Verificar que el pago esté en estado correcto
        if payment.state != 'in_process':
            error_msg = _('El pago %s no está en estado "En Proceso". Estado actual: %s') % (payment.name, payment.state)
            match.state = 'error'
            errors.append(error_msg)
            _logger.warning("✗ ERROR en match %s (Pago: %s): %s", match.id, payment.name, error_msg)
            return
        
        # Buscar la línea del extracto que corresponde a este match
        statement_line = statement.line_ids.filtered(
            lambda l: l.payment_ref == match.extracto_reference and 
                     abs(l.amount - match.extracto_amount) < 0.01
        )
        
        if not statement_line:
            error_msg = _('No se encontró la línea del extracto para el match %s') % match.display_name
            match.state = 'error'
            errors.append(error_msg)
            _logger.warning("✗ ERROR en match %s: %s", match.id, error_msg)
            return
        
        statement_line = statement_line[0]
        
        # VALIDACIÓN ESTRICTA: Verificar que el extracto coincida exactamente con el pago
        # PERO: Si este extracto tiene múltiples pagos, la validación ya se hizo arriba
        extracto_key = (match.extracto_reference or '', match.extracto_amount)
        tiene_multiples_pagos = extracto_key in extractos_con_multiples_pagos
        
        extracto_amount = abs(statement_line.amount)
        payment_amount = abs(payment.amount)
        
        # Solo validar individualmente si NO tiene múltiples pagos
        if not tiene_multiples_pagos and abs(extracto_amount - payment_amount) > 0.01:
            error_msg = _(
                '❌ ERROR: El monto del extracto (%s) no coincide con el monto del pago %s (%s).\n\n'
                'Los montos deben ser EXACTAMENTE iguales para poder reconciliar.\n\n'
                'Extracto: Ref=%s, Monto=%s\n'
                'Pago: %s, Monto=%s'
            ) % (
                extracto_amount, payment.name, payment_amount,
                match.extracto_reference or 'N/A', extracto_amount,
                payment.name, payment_amount
            )
            match.state = 'error'
            errors.append(error_msg)
            _logger.warning("✗ ERROR en match %s: %s", match.id, error_msg)
            return
        
        # Continuar con la reconciliación normal (código del método original)
        if not payment.move_id:
            error_msg = _('El pago %s no tiene un asiento contable asociado.') % payment.name
            match.state = 'error'
            errors.append(error_msg)
            _logger.warning("✗ ERROR en match %s (Pago: %s): %s", match.id, payment.name, error_msg)
            return
        
        # Buscar líneas del pago que son reconciliables
        # Prioridad 1: Cuentas por cobrar/pagar (cuando hay facturas)
        payment_lines = payment.move_id.line_ids.filtered(
            lambda l: l.account_id.account_type in ('asset_receivable', 'liability_payable') and 
                     not l.reconciled and
                     abs(l.amount_residual) > 0.01
        )
        
        # Prioridad 2: Cuentas corrientes reconciliables (cuando hay facturas o pagos directos)
        if not payment_lines:
            payment_lines = payment.move_id.line_ids.filtered(
                lambda l: l.account_id.account_type in ('asset_current', 'liability_current') and
                         l.account_id.reconcile and 
                         not l.reconciled and
                         abs(l.amount_residual) > 0.01
            )
        
        # Prioridad 3: Cuentas bancarias/efectivo del mismo diario (para pagos sin facturas)
        # Esto permite reconciliar pagos directamente con extractos bancarios
        if not payment_lines and statement.journal_id:
            journal_account = statement.journal_id.default_account_id
            if journal_account:
                payment_lines = payment.move_id.line_ids.filtered(
                    lambda l: l.account_id.id == journal_account.id and
                             not l.reconciled and
                             abs(l.amount_residual) > 0.01
                )
        
        # Prioridad 4: Cualquier cuenta reconciliable que no esté reconciliada
        if not payment_lines:
            payment_lines = payment.move_id.line_ids.filtered(
                lambda l: l.account_id.reconcile and 
                         not l.reconciled and
                         abs(l.amount_residual) > 0.01
            )
        
        if not payment_lines:
            error_msg = _('No se encontraron líneas reconciliables para el pago %s') % payment.name
            match.state = 'error'
            errors.append(error_msg)
            _logger.warning("✗ ERROR en match %s (Pago: %s): %s", match.id, payment.name, error_msg)
            return
        
        # Reconciliar usando OCA
        try:
            for payment_line in payment_lines:
                statement_line._add_account_move_line(payment_line, keep_current=False)
            
            statement_line.reconcile_bank_line()
            
            payment.invalidate_recordset(['is_reconciled', 'state'])
            payment = self.env['account.payment'].browse(payment.id)
            
            if payment.is_reconciled and payment.state == 'in_process':
                payment.write({'state': 'paid'})
                _logger.info("Pago %s actualizado a estado 'paid' después de reconciliación", payment.name)
            
            # Contar facturas reconciliadas
            facturas_count = 0
            if payment.partner_id:
                facturas_reconciled = self.env['account.move'].search([
                    ('partner_id', '=', payment.partner_id.id),
                    ('move_type', 'in', ['out_invoice', 'out_refund', 'in_invoice', 'in_refund']),
                    ('state', '=', 'posted'),
                    ('payment_state', '=', 'paid'),
                ])
                for invoice in facturas_reconciled:
                    invoice_lines = invoice.line_ids.filtered(
                        lambda l: l.reconciled and 
                                 l.account_id.account_type in ('asset_receivable', 'liability_payable')
                    )
                    if invoice_lines:
                        payment_lines_reconciled = payment.move_id.line_ids.filtered(
                            lambda l: l.reconciled and l.account_id.account_type in ('asset_receivable', 'liability_payable')
                        )
                        common_reconcile = invoice_lines.mapped('full_reconcile_id') & payment_lines_reconciled.mapped('full_reconcile_id')
                        if common_reconcile:
                            facturas_count += 1
            
            match.state = 'reconciled'
            reconciled_count[0] += 1
            
            if facturas_count > 0:
                _logger.info(
                    "✓ Reconciliación exitosa: Pago %s (monto=%s) reconciliado con extracto %s (monto=%s) y %d factura(s)", 
                    payment.name, payment_amount, statement_line.payment_ref, extracto_amount, facturas_count
                )
            else:
                _logger.info(
                    "✓ Reconciliación exitosa: Pago %s (monto=%s) reconciliado con extracto %s (monto=%s)", 
                    payment.name, payment_amount, statement_line.payment_ref, extracto_amount
                )
                
        except Exception as reconcile_error:
            error_msg = _('Error al reconciliar usando OCA: %s') % str(reconcile_error)
            match.state = 'error'
            errors.append(error_msg)
            _logger.error("Error al reconciliar match %s usando OCA: %s", match.id, str(reconcile_error), exc_info=True)
    
    def _reconcile_multiples_pagos_extracto(self, extracto_matches, statement, extractos_con_multiples_pagos, errors, reconciled_count):
        """Reconciliar múltiples pagos con un solo extracto dividiendo el extracto en múltiples líneas"""
        # Obtener la línea del extracto original (debe ser la misma para todos los matches)
        primer_match = extracto_matches[0]
        original_statement_line = statement.line_ids.filtered(
            lambda l: l.payment_ref == primer_match.extracto_reference and 
                     abs(l.amount - primer_match.extracto_amount) < 0.01
        )
        
        if not original_statement_line:
            error_msg = _('No se encontró la línea del extracto para los matches del extracto %s') % (primer_match.extracto_reference or 'N/A')
            for match in extracto_matches:
                match.state = 'error'
                errors.append(error_msg)
            _logger.warning("✗ ERROR: %s", error_msg)
            return
        
        original_statement_line = original_statement_line[0]
        extracto_amount = abs(original_statement_line.amount)
        
        _logger.info(
            "Reconciliando extracto %s (monto=%s) con %d pagos - Dividiendo extracto en múltiples líneas",
            primer_match.extracto_reference or 'N/A', extracto_amount, len(extracto_matches)
        )
        
        # Verificar que todos los pagos estén en estado correcto antes de proceder
        valid_matches = []
        for match in extracto_matches:
            payment = match.payment_id
            
            if payment.state != 'in_process':
                error_msg = _('El pago %s no está en estado "En Proceso". Estado actual: %s') % (payment.name, payment.state)
                match.state = 'error'
                errors.append(error_msg)
                _logger.warning("✗ ERROR en match %s (Pago: %s): %s", match.id, payment.name, error_msg)
                continue
            
            if not payment.move_id:
                error_msg = _('El pago %s no tiene un asiento contable asociado.') % payment.name
                match.state = 'error'
                errors.append(error_msg)
                _logger.warning("✗ ERROR en match %s (Pago: %s): %s", match.id, payment.name, error_msg)
                continue
            
            valid_matches.append(match)
        
        if not valid_matches:
            _logger.warning("No hay matches válidos para reconciliar extracto %s", primer_match.extracto_reference or 'N/A')
            return
        
        # Dividir el extracto: crear una línea nueva por cada pago
        # IMPORTANTE: OCA no permite múltiples líneas transitorias en el mismo extracto,
        # por lo que debemos crear líneas separadas y reconciliar cada una individualmente
        try:
            # Crear nuevas líneas de extracto para cada pago
            new_statement_lines = []
            for match in valid_matches:
                payment = match.payment_id
                payment_amount = abs(payment.amount)
                
                # Crear una nueva línea de extracto con el monto del pago
                new_line_vals = {
                    'statement_id': statement.id,
                    'date': original_statement_line.date,
                    'payment_ref': original_statement_line.payment_ref,
                    'narration': original_statement_line.narration or '',
                    'amount': -payment_amount if original_statement_line.amount < 0 else payment_amount,
                    'partner_id': original_statement_line.partner_id.id if original_statement_line.partner_id else False,
                }
                
                new_statement_line = self.env['account.bank.statement.line'].create(new_line_vals)
                new_statement_lines.append((new_statement_line, match))
                
                _logger.info(
                    "Creada línea de extracto %s (monto=%s) para pago %s (monto=%s)",
                    new_statement_line.name, new_statement_line.amount, payment.name, payment_amount
                )
            
            # Reconciliar cada pago con su línea de extracto correspondiente
            for new_statement_line, match in new_statement_lines:
                payment = match.payment_id
                
                # Buscar líneas reconciliables del pago
                # Prioridad 1: Cuentas por cobrar/pagar (cuando hay facturas)
                payment_lines = payment.move_id.line_ids.filtered(
                    lambda l: l.account_id.account_type in ('asset_receivable', 'liability_payable') and 
                             not l.reconciled and
                             abs(l.amount_residual) > 0.01
                )
                
                # Prioridad 2: Cuentas corrientes reconciliables (cuando hay facturas o pagos directos)
                if not payment_lines:
                    payment_lines = payment.move_id.line_ids.filtered(
                        lambda l: l.account_id.account_type in ('asset_current', 'liability_current') and
                                 l.account_id.reconcile and 
                                 not l.reconciled and
                                 abs(l.amount_residual) > 0.01
                    )
                
                # Prioridad 3: Cuentas bancarias/efectivo del mismo diario (para pagos sin facturas)
                # Esto permite reconciliar pagos directamente con extractos bancarios
                if not payment_lines and statement.journal_id:
                    journal_account = statement.journal_id.default_account_id
                    if journal_account:
                        payment_lines = payment.move_id.line_ids.filtered(
                            lambda l: l.account_id.id == journal_account.id and
                                     not l.reconciled and
                                     abs(l.amount_residual) > 0.01
                        )
                
                # Prioridad 4: Cualquier cuenta reconciliable que no esté reconciliada
                if not payment_lines:
                    payment_lines = payment.move_id.line_ids.filtered(
                        lambda l: l.account_id.reconcile and 
                                 not l.reconciled and
                                 abs(l.amount_residual) > 0.01
                    )
                
                if not payment_lines:
                    error_msg = _('No se encontraron líneas reconciliables para el pago %s') % payment.name
                    match.state = 'error'
                    errors.append(error_msg)
                    _logger.warning("✗ ERROR en match %s (Pago: %s): %s", match.id, payment.name, error_msg)
                    continue
                
                # Agregar la línea del pago al extracto
                for payment_line in payment_lines:
                    new_statement_line._add_account_move_line(payment_line, keep_current=False)
                
                # Reconciliar usando OCA
                new_statement_line.reconcile_bank_line()
                
                # Marcar el match como reconciliado
                payment.invalidate_recordset(['is_reconciled', 'state'])
                payment = self.env['account.payment'].browse(payment.id)
                
                if payment.is_reconciled and payment.state == 'in_process':
                    payment.write({'state': 'paid'})
                    _logger.info("Pago %s actualizado a estado 'paid' después de reconciliación", payment.name)
                
                match.state = 'reconciled'
                reconciled_count[0] += 1
                
                _logger.info(
                    "✓ Reconciliación exitosa: Pago %s (monto=%s) reconciliado con línea de extracto %s (monto=%s)",
                    payment.name, abs(payment.amount), new_statement_line.name, new_statement_line.amount
                )
            
            # NO eliminar la línea original del extracto
            # Mantenerla para referencia, pero las líneas creadas son las que se usan para reconciliar
            # Esto permite que al romper la conciliación, simplemente desreconciliemos cada línea
            # sin necesidad de consolidar de vuelta
            _logger.info("Línea original del extracto %s mantenida. Se crearon %d líneas adicionales para reconciliación", 
                       primer_match.extracto_reference or 'N/A', len(new_statement_lines))
            
            _logger.info(
                "✓ Reconciliación exitosa: Extracto %s (monto=%s) dividido y reconciliado con %d pago(s)",
                primer_match.extracto_reference or 'N/A', extracto_amount, len(valid_matches)
            )
            
        except Exception as reconcile_error:
            error_msg = _('Error al reconciliar extracto %s con múltiples pagos: %s') % (primer_match.extracto_reference or 'N/A', str(reconcile_error))
            for match in valid_matches:
                if match.state != 'reconciled':  # Solo marcar como error si no estaba reconciliado
                    match.state = 'error'
                    errors.append(error_msg)
            _logger.error("Error al reconciliar extracto %s: %s", primer_match.extracto_reference or 'N/A', str(reconcile_error), exc_info=True)
    
    def action_view_statement(self):
        """Abrir el extracto bancario creado"""
        self.ensure_one()
        if not self.statement_id:
            raise UserError(_('No se ha creado ningún extracto bancario para este registro.'))
        
        return {
            'type': 'ir.actions.act_window',
            'name': _('Extracto Bancario'),
            'res_model': 'account.bank.statement',
            'res_id': self.statement_id.id,
            'view_mode': 'form',
            'target': 'current',
        }
    
    def action_unreconcile(self):
        """Deshacer todas las conciliaciones de este registro y volver a borrador"""
        self.ensure_one()
        
        if self.state != 'reconciled':
            raise UserError(_('Solo se pueden deshacer conciliaciones en estado "Reconciliado".'))
        
        # Obtener todos los matches reconciliados
        reconciled_matches = self.match_line_ids.filtered(lambda m: m.state == 'reconciled')
        
        if not reconciled_matches:
            raise UserError(_('No hay conciliaciones para deshacer.'))
        
        unreconciled_count = 0
        errors = []
        
        # Agrupar matches por extracto para manejar extractos divididos
        matches_by_extracto = {}
        for match in reconciled_matches:
            extracto_key = (match.extracto_reference or '', match.extracto_amount)
            if extracto_key not in matches_by_extracto:
                matches_by_extracto[extracto_key] = []
            matches_by_extracto[extracto_key].append(match)
        
        # Procesar cada extracto
        for extracto_key, extracto_matches in matches_by_extracto.items():
            extracto_ref, extracto_amount = extracto_key
            
            # Buscar todas las líneas del extracto que coinciden con esta referencia
            # (pueden ser líneas divididas con diferentes montos)
            statement_lines = self.statement_id.line_ids.filtered(
                lambda l: l.payment_ref == extracto_ref
            )
            
            # Si hay múltiples matches para este extracto, probablemente fue dividido
            is_divided_extract = len(extracto_matches) > 1
            
            if is_divided_extract:
                _logger.info(
                    "Extracto %s (monto original=%s) fue dividido en %d líneas. Consolidando...",
                    extracto_ref or 'N/A', extracto_amount, len(statement_lines)
                )
            
            # Procesar cada match de este extracto
            for match in extracto_matches:
                try:
                    payment = match.payment_id
                    
                    if not payment.move_id:
                        match.state = 'pending'
                        continue
                    
                    # IMPORTANTE: Solo desreconciliar las líneas relacionadas con el extracto bancario,
                    # NO las líneas de facturas. Esto mantiene la relación pago-factura.
                    
                    # Buscar la línea del extracto que corresponde a este pago
                    # Estrategia: primero buscar por reconciliación, luego por monto
                    statement_line = self.env['account.bank.statement.line']
                    
                    # PRIMERO: Buscar líneas del extracto que estén reconciliadas con este pago específico
                    # Esto funciona incluso si hay facturas en la reconciliación
                    for sl in statement_lines:
                        if sl.move_id:
                            # Buscar líneas del extracto reconciliadas
                            sl_reconciled_lines = sl.move_id.line_ids.filtered(lambda l: l.reconciled)
                            for sl_line in sl_reconciled_lines:
                                if sl_line.full_reconcile_id:
                                    # Verificar si esta reconciliación incluye el pago
                                    reconciled_payment_lines = sl_line.full_reconcile_id.reconciled_line_ids.filtered(
                                        lambda l: l.move_id == payment.move_id
                                    )
                                    if reconciled_payment_lines:
                                        # Esta línea del extracto está reconciliada con este pago
                                        statement_line = sl
                                        _logger.info("Encontrada línea de extracto %s reconciliada con pago %s", sl.name, payment.name)
                                        break
                            if statement_line:
                                break
                    
                    # SEGUNDO: Si no encontramos por reconciliación, buscar por monto
                    if not statement_line:
                        if is_divided_extract:
                            # Buscar la línea del extracto que tiene el mismo monto que el pago
                            # Excluir la línea original (que tiene el monto total)
                            statement_line = statement_lines.filtered(
                                lambda l: abs(abs(l.amount) - abs(payment.amount)) < 0.01 and
                                         abs(l.amount) != extracto_amount  # Excluir la línea original
                            )
                            # Si no encontramos por monto exacto, buscar cualquier línea dividida (que no sea la original)
                            if not statement_line:
                                statement_line = statement_lines.filtered(
                                    lambda l: abs(l.amount) != extracto_amount
                                )[:1] if statement_lines else self.env['account.bank.statement.line']
                        else:
                            # Buscar la línea del extracto por monto original
                            statement_line = statement_lines.filtered(
                                lambda l: abs(l.amount - extracto_amount) < 0.01
                            )
                        
                        if not statement_line:
                            # Si no encontramos la línea exacta, buscar cualquier línea con la misma referencia
                            if is_divided_extract:
                                statement_line = statement_lines.filtered(
                                    lambda l: abs(l.amount) != extracto_amount
                                )[:1] if statement_lines else self.env['account.bank.statement.line']
                            else:
                                statement_line = statement_lines[:1] if statement_lines else self.env['account.bank.statement.line']
                    
                    # Obtener las líneas reconciliadas del pago
                    payment_lines_reconciled = payment.move_id.line_ids.filtered(
                        lambda l: l.reconciled
                    )
                    
                    # Identificar qué reconciliaciones incluyen el extracto bancario
                    # IMPORTANTE: Solo desreconciliar las líneas relacionadas con el extracto,
                    # NO las líneas relacionadas solo con facturas
                    lines_to_unreconcile = self.env['account.move.line']
                    extracto_reconcile_ids = set()
                    
                    if statement_line and statement_line[0].move_id:
                        statement_line = statement_line[0]
                        # Obtener las líneas del asiento del extracto que están reconciliadas
                        extracto_lines = statement_line.move_id.line_ids.filtered(
                            lambda l: l.reconciled
                        )
                        
                        # Identificar todas las reconciliaciones que incluyen el extracto
                        for extracto_line in extracto_lines:
                            if extracto_line.full_reconcile_id:
                                extracto_reconcile_ids.add(extracto_line.full_reconcile_id.id)
                        
                        # Ahora, para cada reconciliación que incluye el extracto,
                        # identificar las líneas que podemos desreconciliar (solo extracto y pago, NO facturas)
                        for reconcile_id in extracto_reconcile_ids:
                            reconcile = self.env['account.full.reconcile'].browse(reconcile_id)
                            all_reconciled = reconcile.reconciled_line_ids
                            
                            # Separar: líneas del extracto, líneas del pago, y líneas de facturas
                            extracto_lines_in_reconcile = all_reconciled.filtered(
                                lambda l: l.move_id == statement_line.move_id
                            )
                            payment_lines_in_reconcile = all_reconciled.filtered(
                                lambda l: l.move_id == payment.move_id and 
                                         l.move_id.move_type == 'entry'
                            )
                            invoice_lines_in_reconcile = all_reconciled.filtered(
                                lambda l: l.move_id.move_type in ['out_invoice', 'in_invoice', 'out_refund', 'in_refund']
                            )
                            
                            # Si hay facturas en esta reconciliación, NO podemos desreconciliar sin afectar la factura
                            # PERO: Si el pago tiene múltiples reconciliaciones (una con extracto, otra con factura),
                            # podemos intentar desreconciliar solo la parte del extracto
                            # Por ahora, si hay facturas, NO desreconciliamos para mantener la relación pago-factura
                            if invoice_lines_in_reconcile:
                                _logger.info("Pago %s tiene facturas reconciliadas en esta reconciliación con extracto. No se puede desreconciliar sin afectar la factura. Se mantiene la relación pago-factura.", payment.name)
                                # NO agregamos estas líneas a lines_to_unreconcile
                            else:
                                # Solo desreconciliar extracto y pago si NO hay facturas
                                lines_to_unreconcile |= extracto_lines_in_reconcile
                                lines_to_unreconcile |= payment_lines_in_reconcile
                    
                    # Si no encontramos líneas del extracto, buscar reconciliaciones del pago que NO incluyan facturas
                    if not lines_to_unreconcile and payment_lines_reconciled and statement_line:
                        statement_line_obj = statement_line[0] if isinstance(statement_line, type(self.env['account.bank.statement.line'])) else statement_line
                        extracto_move_id = statement_line_obj.move_id.id if statement_line_obj.move_id else None
                        
                        # Buscar reconciliaciones del pago que NO tengan facturas pero sí tengan extracto
                        for line in payment_lines_reconciled:
                            if line.full_reconcile_id:
                                reconciled_moves = line.full_reconcile_id.reconciled_line_ids.mapped('move_id')
                                has_invoices = reconciled_moves.filtered(
                                    lambda m: m.move_type in ['out_invoice', 'in_invoice', 'out_refund', 'in_refund']
                                )
                                
                                # Solo agregar si NO hay facturas y si esta reconciliación NO está ya en extracto_reconcile_ids
                                if not has_invoices and line.full_reconcile_id.id not in extracto_reconcile_ids:
                                    # Verificar si esta reconciliación incluye alguna línea del extracto
                                    all_lines = line.full_reconcile_id.reconciled_line_ids
                                    has_extracto = False
                                    if extracto_move_id:
                                        has_extracto = any(
                                            l.move_id.id == extracto_move_id 
                                            for l in all_lines
                                        )
                                    
                                    # Si incluye extracto, agregar las líneas
                                    if has_extracto:
                                        lines_to_unreconcile |= all_lines.filtered(
                                            lambda l: l.move_id.move_type == 'entry'
                                        )
                    
                    desreconciliado_exitoso = False
                    if lines_to_unreconcile:
                        try:
                            # Desreconciliar solo las líneas del extracto/pago (no facturas)
                            lines_to_unreconcile.remove_move_reconcile()
                            unreconciled_count += 1
                            desreconciliado_exitoso = True
                            _logger.info("Desreconciliación exitosa (solo extracto): Pago %s. Se mantiene la relación con facturas.", payment.name)
                        except Exception as e:
                            # Error al desreconciliar, pero continuamos con las demás
                            _logger.warning("⚠️ No se pudo desreconciliar pago %s: %s. Se dejará para resolución manual.", payment.name, str(e))
                            errors.append(_('⚠️ Pago %s: %s (dejar para resolución manual)') % (payment.name, str(e)))
                            # Marcar el match como que necesita atención manual
                            match.state = 'reconciled'  # Mantener en reconciliado para que se vea que necesita atención
                            continue  # Continuar con el siguiente match
                    
                    # Solo actualizar el estado si la desreconciliación fue exitosa
                    if desreconciliado_exitoso:
                        # Actualizar el estado del match a 'pending' (no reconciliado)
                        match.state = 'pending'
                        
                        # Recalcular is_reconciled del pago
                        payment.invalidate_recordset(['is_reconciled', 'state'])
                        payment = self.env['account.payment'].browse(payment.id)
                        
                        # Si el pago ya no está reconciliado pero estaba en 'paid', volver a 'in_process'
                        # Esto mantiene el pago "en proceso de pago" como mencionó el usuario
                        if not payment.is_reconciled and payment.state == 'paid':
                            payment.write({'state': 'in_process'})
                            _logger.info("Pago %s vuelto a estado 'in_process' después de desreconciliación", payment.name)
                        elif payment.state == 'paid' and payment.is_reconciled:
                            # Si el pago sigue reconciliado (por ejemplo, con facturas), mantenerlo en 'paid'
                            _logger.info("Pago %s sigue reconciliado (probablemente con facturas), manteniendo estado 'paid'", payment.name)
                    else:
                        # Si no había líneas para desreconciliar, investigar por qué
                        if not lines_to_unreconcile:
                            # Verificar si el pago está reconciliado solo con facturas (no con extracto)
                            payment_has_invoices_only = False
                            if payment_lines_reconciled:
                                for line in payment_lines_reconciled:
                                    if line.full_reconcile_id:
                                        reconciled_moves = line.full_reconcile_id.reconciled_line_ids.mapped('move_id')
                                        has_invoices = reconciled_moves.filtered(
                                            lambda m: m.move_type in ['out_invoice', 'in_invoice', 'out_refund', 'in_refund']
                                        )
                                        has_extracto = False
                                        if statement_line and statement_line.move_id:
                                            has_extracto = any(
                                                l.move_id == statement_line.move_id 
                                                for l in line.full_reconcile_id.reconciled_line_ids
                                            )
                                        
                                        if has_invoices and not has_extracto:
                                            payment_has_invoices_only = True
                                            break
                            
                            if payment_has_invoices_only:
                                _logger.info("Pago %s está reconciliado solo con facturas (no con extracto). No hay nada que desreconciliar del extracto.", payment.name)
                                # Marcar como pending porque no hay relación con extracto que deshacer
                                match.state = 'pending'
                            else:
                                _logger.warning("⚠️ Pago %s no tiene líneas reconciliables con extracto o la reconciliación incluye facturas que no se pueden separar. Se dejará para resolución manual.", payment.name)
                                match.state = 'reconciled'  # Mantener en reconciliado
                                errors.append(_('⚠️ Pago %s: No se puede desreconciliar del extracto sin afectar facturas. Requiere resolución manual.') % payment.name)
                    
                except Exception as e:
                    # Error general, pero continuamos con las demás conciliaciones
                    _logger.error("⚠️ Error al procesar match %s (Pago: %s): %s. Se dejará para resolución manual.", 
                                match.id, match.payment_id.name if match.payment_id else 'N/A', str(e), exc_info=True)
                    errors.append(_('⚠️ Error en match %s (Pago: %s): %s (dejar para resolución manual)') % 
                                (match.id, match.payment_id.name if match.payment_id else 'N/A', str(e)))
                    # Mantener el match en estado reconciliado para que se vea que necesita atención
                    match.state = 'reconciled'
                    continue  # Continuar con el siguiente match
            
            # NO consolidar las líneas divididas de vuelta
            # Las líneas divididas se mantienen como están
            # Esto evita problemas con asientos contables reconciliados
            # Cada línea puede ser desreconciliada individualmente sin afectar a las demás
            if is_divided_extract:
                _logger.info(
                    "Extracto %s mantiene sus líneas divididas. No se consolida para evitar problemas con asientos reconciliados.",
                    extracto_ref or 'N/A'
                )
        
        # Actualizar estado del registro
        remaining_reconciled = self.match_line_ids.filtered(lambda m: m.state == 'reconciled')
        
        # Actualizar el campo de errores si los hay
        if errors:
            error_message = _('Errores encontrados al romper conciliaciones:\n\n') + '\n'.join(errors[:20])  # Limitar a 20 errores
            if len(errors) > 20:
                error_message += _('\n\n... y %d error(es) más') % (len(errors) - 20)
            self.unreconcile_error_message = error_message
        else:
            self.unreconcile_error_message = False
        
        if not remaining_reconciled:
            self.state = 'matches_found'
            _logger.info("✓ Todos los matches fueron desreconciliados exitosamente. Estado cambiado a 'matches_found'")
        else:
            _logger.warning("⚠️ %d match(es) no pudieron ser desreconciliados. Estado permanece en 'reconciled'", len(remaining_reconciled))
        
        # Solo cerrar la acción sin recargar - la vista se actualizará automáticamente
        return {'type': 'ir.actions.act_window_close'}


class MatchRegisterConciliacionLine(models.Model):
    _name = 'match.register.conciliacion.line'
    _description = 'Línea de Match de Conciliación'
    _order = 'date desc, amount desc'

    match_register_id = fields.Many2one(
        'match.register.conciliacion',
        string='Registro de Conciliación',
        required=True,
        ondelete='cascade',
        index=True,
    )
    
    payment_id = fields.Many2one(
        'account.payment',
        string='Pago',
        required=True,
        ondelete='cascade',
        index=True,
    )
    
    # Información del extracto (datos del Excel/CSV)
    extracto_date = fields.Date(
        string='Fecha Extracto',
        help='Fecha de la línea del extracto bancario',
    )
    
    extracto_reference = fields.Char(
        string='Referencia Extracto',
        help='Referencia de la línea del extracto bancario',
    )
    
    extracto_description = fields.Char(
        string='Descripción Extracto',
        help='Descripción de la línea del extracto bancario',
    )
    
    extracto_amount = fields.Monetary(
        string='Monto Extracto',
        help='Monto de la línea del extracto bancario',
    )
    
    extracto_partner_id = fields.Many2one(
        'res.partner',
        string='Cliente/Proveedor Extracto',
        help='Cliente/Proveedor de la línea del extracto bancario',
    )
    
    # Información del pago
    date = fields.Date(
        string='Fecha Pago',
        related='payment_id.date',
        store=True,
        readonly=True,
    )
    
    amount = fields.Monetary(
        string='Monto Pago',
        related='payment_id.amount',
        store=True,
        readonly=True,
    )
    
    currency_id = fields.Many2one(
        'res.currency',
        string='Moneda',
        related='payment_id.currency_id',
        store=True,
        readonly=True,
    )
    
    partner_id = fields.Many2one(
        'res.partner',
        string='Cliente/Proveedor',
        related='payment_id.partner_id',
        store=True,
        readonly=True,
    )
    
    payment_memo = fields.Char(
        string='Memo del Pago',
        related='payment_id.memo',
        store=True,
        readonly=True,
    )
    
    payment_name = fields.Char(
        string='Nombre del Pago',
        related='payment_id.name',
        store=True,
        readonly=True,
    )
    
    # Estado y control
    state = fields.Selection([
        ('pending', 'Pendiente'),
        ('reconciled', 'Reconciliado'),
        ('error', 'Error'),
    ], string='Estado', default='pending', required=True, index=True)
    
    selected = fields.Boolean(
        string='Seleccionado',
        default=True,
        help='Marcar para incluir este match en la reconciliación',
    )
    
    match_score = fields.Float(
        string='Puntuación de Match',
        compute='_compute_match_score',
        store=True,
        help='Puntuación que indica qué tan probable es que sea un match correcto (0-100)',
    )
    
    company_id = fields.Many2one(
        'res.company',
        string='Compañía',
        related='match_register_id.company_id',
        store=True,
        readonly=True,
    )
    
    @api.depends('extracto_amount', 'amount', 'extracto_reference', 'payment_memo', 
                 'extracto_partner_id', 'partner_id', 'extracto_date', 'date')
    def _compute_match_score(self):
        """Calcular puntuación de match - Mayor puntuación = mejor match"""
        for record in self:
            score = 0.0
            
            # Match por monto (40 puntos máximo)
            if record.extracto_amount and record.amount:
                if abs(record.extracto_amount - abs(record.amount)) < 0.01:
                    score += 40  # Monto exacto
                elif abs(record.extracto_amount - abs(record.amount)) / max(abs(record.extracto_amount), 1) < 0.01:
                    score += 30  # Monto muy cercano
                elif abs(record.extracto_amount - abs(record.amount)) / max(abs(record.extracto_amount), 1) < 0.05:
                    score += 20  # Monto cercano (5% tolerancia)
            
            # Match por partner (30 puntos máximo) - MUY IMPORTANTE
            if record.extracto_partner_id and record.partner_id:
                if record.extracto_partner_id == record.partner_id:
                    score += 30
            
            # Match por referencia (25 puntos máximo) - IMPORTANTE para distinguir facturas
            if record.extracto_reference and record.payment_memo:
                ref_lower = record.extracto_reference.lower().strip()
                pay_ref_lower = record.payment_memo.lower().strip()
                if ref_lower == pay_ref_lower:
                    score += 25  # Referencia exacta
                elif ref_lower in pay_ref_lower or pay_ref_lower in ref_lower:
                    score += 15  # Referencia parcial
            
            # Match por fecha (5 puntos máximo) - Ayuda a distinguir cuando hay múltiples facturas
            if record.extracto_date and record.date:
                days_diff = abs((record.extracto_date - record.date).days)
                if days_diff == 0:
                    score += 5  # Misma fecha
                elif days_diff <= 7:
                    score += 3  # Dentro de una semana
                elif days_diff <= 30:
                    score += 1  # Dentro de un mes
            
            record.match_score = min(score, 100.0)
