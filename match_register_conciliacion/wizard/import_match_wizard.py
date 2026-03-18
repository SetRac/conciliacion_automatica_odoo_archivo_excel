# -*- coding: utf-8 -*-

from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError
from datetime import datetime
import base64
import io
import logging

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import csv
except ImportError:
    csv = None

_logger = logging.getLogger(__name__)


class ImportMatchWizard(models.TransientModel):
    _name = 'import.match.wizard'
    _description = 'Wizard para Importar Extracto y Buscar Matches'

    journal_id = fields.Many2one(
        'account.journal',
        string='Diario',
        required=True,
        domain=[('type', '=', 'bank')],
        help='Seleccione el diario bancario para el extracto',
    )
    
    file = fields.Binary(
        string='Archivo Excel/CSV',
        required=True,
        help='Suba el archivo Excel o CSV con el extracto bancario',
    )
    
    filename = fields.Char(string='Nombre del Archivo')
    
    file_type = fields.Selection([
        ('excel', 'Excel (.xlsx)'),
        ('csv', 'CSV'),
    ], string='Tipo de Archivo', default='excel', required=True)
    
    # Mapeo de columnas
    column_date = fields.Char(
        string='Columna Fecha',
        default='A',
        help='Columna del Excel que contiene la fecha (ej: A, B, C...)',
    )
    
    column_reference = fields.Char(
        string='Columna Referencia',
        default='B',
        help='Columna del Excel que contiene la referencia/número de operación',
    )
    
    column_partner = fields.Char(
        string='Columna Cliente/Proveedor',
        default='C',
        help='Columna del Excel que contiene el nombre del cliente/proveedor',
    )
    
    column_amount = fields.Char(
        string='Columna Monto',
        default='D',
        help='Columna del Excel que contiene el monto',
    )
    
    column_description = fields.Char(
        string='Columna Descripción',
        default='E',
        help='Columna del Excel que contiene la descripción/glosa',
    )
    
    column_vat = fields.Char(
        string='Columna RIF/NIT',
        default='',
        help='Columna del Excel que contiene el RIF/NIT (opcional)',
    )
    
    # Configuración adicional
    start_row = fields.Integer(
        string='Fila de Inicio',
        default=2,
        help='Número de fila donde comienzan los datos (1 = primera fila)',
    )
    
    date_format = fields.Char(
        string='Formato de Fecha',
        default='%d/%m/%Y',
        help='Formato de fecha en el Excel (ej: %d/%m/%Y, %Y-%m-%d)',
    )
    
    # Información del mapeo
    mapping_info = fields.Html(
        string='Información del Mapeo',
        compute='_compute_mapping_info',
        readonly=True,
    )
    
    @api.depends('column_date', 'column_reference', 'column_partner', 'column_amount', 
                 'column_description', 'column_vat', 'start_row', 'date_format')
    def _compute_mapping_info(self):
        """Mostrar información del mapeo de columnas"""
        for record in self:
            info = f"""
            <div class="alert alert-info">
                <h5><strong>Mapeo de Columnas:</strong></h5>
                <table class="table table-sm">
                    <thead>
                        <tr>
                            <th>Campo Odoo</th>
                            <th>Columna Excel/CSV</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr>
                            <td><strong>Fecha</strong></td>
                            <td>Columna <strong>{record.column_date or 'No definida'}</strong></td>
                        </tr>
                        <tr>
                            <td><strong>Referencia</strong></td>
                            <td>Columna <strong>{record.column_reference or 'No definida'}</strong></td>
                        </tr>
                        <tr>
                            <td><strong>Cliente/Proveedor</strong></td>
                            <td>Columna <strong>{record.column_partner or 'No definida'}</strong></td>
                        </tr>
                        <tr>
                            <td><strong>Monto</strong></td>
                            <td>Columna <strong>{record.column_amount or 'No definida'}</strong></td>
                        </tr>
                        <tr>
                            <td><strong>Descripción</strong></td>
                            <td>Columna <strong>{record.column_description or 'No definida'}</strong></td>
                        </tr>
                        <tr>
                            <td><strong>RIF/NIT</strong></td>
                            <td>Columna <strong>{record.column_vat or 'No definida (opcional)'}</strong></td>
                        </tr>
                    </tbody>
                </table>
                <p><strong>Fila de inicio:</strong> {record.start_row}</p>
                <p><strong>Formato de fecha:</strong> {record.date_format}</p>
            </div>
            """
            record.mapping_info = info
    
    def _column_to_index(self, column):
        """Convertir letra de columna a índice (A=0, B=1, etc.)"""
        if not column:
            return None
        column = column.upper().strip()
        index = 0
        for char in column:
            index = index * 26 + (ord(char) - ord('A') + 1)
        return index - 1
    
    def _parse_date(self, date_str, date_format):
        """Parsear fecha según el formato especificado"""
        if not date_str:
            return None
        
        # Intentar diferentes formatos comunes
        formats_to_try = [
            date_format,
            '%d/%m/%Y',
            '%Y-%m-%d',
            '%d-%m-%Y',
            '%Y/%m/%d',
            '%d.%m.%Y',
        ]
        
        for fmt in formats_to_try:
            try:
                return datetime.strptime(str(date_str).strip(), fmt).date()
            except (ValueError, AttributeError):
                continue
        
        # Si es un número (fecha serial de Excel)
        try:
            from datetime import datetime, timedelta
            excel_epoch = datetime(1899, 12, 30)
            days = int(float(date_str))
            return (excel_epoch + timedelta(days=days)).date()
        except (ValueError, TypeError):
            pass
        
        raise UserError(_('No se pudo parsear la fecha: %s con formato: %s') % (date_str, date_format))
    
    def action_load_and_find_matches(self):
        """Cargar extracto y buscar matches automáticamente"""
        self.ensure_one()
        
        if not self.file:
            raise UserError(_('Por favor suba un archivo Excel o CSV.'))
        
        if not self.journal_id:
            raise UserError(_('Por favor seleccione un diario.'))
        
        # Crear registro de conciliación
        match_register = self.env['match.register.conciliacion'].create({
            'journal_id': self.journal_id.id,
            'date': fields.Date.today(),
        })
        
        # Procesar archivo
        if self.file_type == 'excel':
            if not openpyxl:
                raise UserError(_(
                    'La librería openpyxl no está instalada. '
                    'Por favor instálela con: pip install openpyxl'
                ))
            lines_data = self._process_excel()
        else:
            if not csv:
                raise UserError(_('El módulo csv no está disponible.'))
            lines_data = self._process_csv()
        
        # Buscar matches para cada línea del extracto
        matches_created = []
        for line_data in lines_data:
            # Buscar pagos que coincidan
            domain = [
                ('journal_id', '=', self.journal_id.id),
                ('state', '=', 'posted'),
                ('is_reconciled', '=', False),
            ]
            
            # Buscar por monto
            matching_payments = self.env['account.payment'].search(domain).filtered(
                lambda p: abs(p.amount - abs(line_data['amount'])) < 0.01
            )
            
            # Si no hay match por monto exacto, buscar por monto aproximado (5% tolerancia)
            if not matching_payments:
                matching_payments = self.env['account.payment'].search(domain).filtered(
                    lambda p: abs(p.amount - abs(line_data['amount'])) / max(abs(line_data['amount']), 1) < 0.05
                )
            
            # Si no hay match por monto, buscar por referencia
            if not matching_payments and line_data.get('reference'):
                matching_payments = self.env['account.payment'].search(domain).filtered(
                    lambda p: line_data['reference'].lower() in (p.memo or '').lower() or 
                             (p.memo or '').lower() in line_data['reference'].lower()
                )
            
            # Si hay partner, filtrar por partner
            if matching_payments and line_data.get('partner_id'):
                matching_payments = matching_payments.filtered(
                    lambda p: p.partner_id.id == line_data['partner_id']
                )
            
            # Crear match para cada pago encontrado
            for payment in matching_payments:
                match_vals = {
                    'match_register_id': match_register.id,
                    'payment_id': payment.id,
                    'extracto_date': line_data.get('date'),
                    'extracto_reference': line_data.get('reference', ''),
                    'extracto_amount': line_data.get('amount'),
                    'extracto_partner_id': line_data.get('partner_id', False),
                    'selected': True,
                }
                match = self.env['match.register.conciliacion.line'].create(match_vals)
                matches_created.append(match)
        
        if matches_created:
            match_register.state = 'matches_found'
        else:
            match_register.state = 'draft'
        
        return {
            'type': 'ir.actions.act_window',
            'name': _('Revisar Matches'),
            'res_model': 'match.register.conciliacion',
            'res_id': match_register.id,
            'view_mode': 'form',
            'target': 'current',
        }
    
    def _process_excel(self):
        """Procesar archivo Excel"""
        try:
            file_data = base64.b64decode(self.file)
            workbook = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError(_('Error al leer el archivo Excel: %s') % str(e))
        
        # Obtener índices de columnas
        col_date_idx = self._column_to_index(self.column_date)
        col_ref_idx = self._column_to_index(self.column_reference)
        col_partner_idx = self._column_to_index(self.column_partner)
        col_amount_idx = self._column_to_index(self.column_amount)
        col_desc_idx = self._column_to_index(self.column_description)
        col_vat_idx = self._column_to_index(self.column_vat) if self.column_vat else None
        
        if col_date_idx is None or col_ref_idx is None or col_partner_idx is None or col_amount_idx is None:
            raise UserError(_('Debe definir al menos las columnas: Fecha, Referencia, Cliente/Proveedor y Monto.'))
        
        lines_data = []
        errors = []
        start_row = max(1, self.start_row)
        
        for row_idx in range(start_row, sheet.max_row + 1):
            try:
                # Leer valores
                date_val = sheet.cell(row=row_idx, column=col_date_idx + 1).value
                ref_val = sheet.cell(row=row_idx, column=col_ref_idx + 1).value
                partner_val = sheet.cell(row=row_idx, column=col_partner_idx + 1).value
                amount_val = sheet.cell(row=row_idx, column=col_amount_idx + 1).value
                desc_val = sheet.cell(row=row_idx, column=col_desc_idx + 1).value if col_desc_idx is not None else ''
                vat_val = sheet.cell(row=row_idx, column=col_vat_idx + 1).value if col_vat_idx is not None else ''
                
                if not date_val or not ref_val or amount_val is None:
                    continue
                
                # Parsear fecha
                try:
                    date = self._parse_date(date_val, self.date_format)
                except Exception as e:
                    errors.append(_('Fila %d: Error en fecha: %s') % (row_idx, str(e)))
                    continue
                
                # Parsear monto
                try:
                    amount = float(amount_val)
                except (ValueError, TypeError):
                    errors.append(_('Fila %d: Error en monto: %s') % (row_idx, str(amount_val)))
                    continue
                
                if amount == 0:
                    continue
                
                # Buscar partner
                partner_id = False
                if partner_val:
                    partner = self.env['res.partner'].search([
                        '|', '|',
                        ('name', 'ilike', str(partner_val).strip()),
                        ('vat', '=', str(vat_val).strip()) if vat_val else (1, '=', 0),
                        ('ref', '=', str(ref_val).strip()),
                    ], limit=1)
                    
                    if not partner and vat_val:
                        partner = self.env['res.partner'].search([
                            ('vat', '=', str(vat_val).strip())
                        ], limit=1)
                    
                    if partner:
                        partner_id = partner.id
                
                lines_data.append({
                    'date': date,
                    'reference': str(ref_val).strip() if ref_val else '',
                    'partner_id': partner_id,
                    'amount': amount,
                    'description': str(desc_val).strip() if desc_val else '',
                })
                
            except Exception as e:
                errors.append(_('Fila %d: %s') % (row_idx, str(e)))
                _logger.error("Error procesando fila %d: %s", row_idx, str(e), exc_info=True)
        
        if errors:
            _logger.warning("Errores al procesar Excel: %s", '\n'.join(errors[:10]))
        
        return lines_data
    
    def _process_csv(self):
        """Procesar archivo CSV"""
        try:
            file_data = base64.b64decode(self.file)
            content = file_data.decode('utf-8')
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)
        except Exception as e:
            raise UserError(_('Error al leer el archivo CSV: %s') % str(e))
        
        # Obtener índices de columnas (en CSV son números, no letras)
        try:
            col_date_idx = int(self.column_date) - 1 if self.column_date.isdigit() else self._column_to_index(self.column_date)
            col_ref_idx = int(self.column_reference) - 1 if self.column_reference.isdigit() else self._column_to_index(self.column_reference)
            col_partner_idx = int(self.column_partner) - 1 if self.column_partner.isdigit() else self._column_to_index(self.column_partner)
            col_amount_idx = int(self.column_amount) - 1 if self.column_amount.isdigit() else self._column_to_index(self.column_amount)
            col_desc_idx = int(self.column_description) - 1 if self.column_description.isdigit() else self._column_to_index(self.column_description)
            col_vat_idx = int(self.column_vat) - 1 if self.column_vat and self.column_vat.isdigit() else (self._column_to_index(self.column_vat) if self.column_vat else None)
        except (ValueError, AttributeError):
            raise UserError(_('Las columnas deben ser números (1, 2, 3...) para archivos CSV.'))
        
        if col_date_idx is None or col_ref_idx is None or col_partner_idx is None or col_amount_idx is None:
            raise UserError(_('Debe definir al menos las columnas: Fecha, Referencia, Cliente/Proveedor y Monto.'))
        
        lines_data = []
        errors = []
        start_row = max(0, self.start_row - 1)  # CSV es 0-indexed
        
        for row_idx, row in enumerate(rows[start_row:], start=start_row + 1):
            try:
                if len(row) <= max(col_date_idx, col_ref_idx, col_partner_idx, col_amount_idx):
                    continue
                
                date_val = row[col_date_idx] if col_date_idx < len(row) else None
                ref_val = row[col_ref_idx] if col_ref_idx < len(row) else None
                partner_val = row[col_partner_idx] if col_partner_idx < len(row) else None
                amount_val = row[col_amount_idx] if col_amount_idx < len(row) else None
                desc_val = row[col_desc_idx] if col_desc_idx is not None and col_desc_idx < len(row) else ''
                vat_val = row[col_vat_idx] if col_vat_idx is not None and col_vat_idx < len(row) else ''
                
                if not date_val or not ref_val or not amount_val:
                    continue
                
                # Parsear fecha
                try:
                    date = self._parse_date(date_val, self.date_format)
                except Exception as e:
                    errors.append(_('Fila %d: Error en fecha: %s') % (row_idx, str(e)))
                    continue
                
                # Parsear monto
                try:
                    amount = float(amount_val)
                except (ValueError, TypeError):
                    errors.append(_('Fila %d: Error en monto: %s') % (row_idx, str(amount_val)))
                    continue
                
                if amount == 0:
                    continue
                
                # Buscar partner
                partner_id = False
                if partner_val:
                    partner = self.env['res.partner'].search([
                        '|', '|',
                        ('name', 'ilike', str(partner_val).strip()),
                        ('vat', '=', str(vat_val).strip()) if vat_val else (1, '=', 0),
                        ('ref', '=', str(ref_val).strip()),
                    ], limit=1)
                    
                    if not partner and vat_val:
                        partner = self.env['res.partner'].search([
                            ('vat', '=', str(vat_val).strip())
                        ], limit=1)
                    
                    if partner:
                        partner_id = partner.id
                
                lines_data.append({
                    'date': date,
                    'reference': str(ref_val).strip() if ref_val else '',
                    'partner_id': partner_id,
                    'amount': amount,
                    'description': str(desc_val).strip() if desc_val else '',
                })
                
            except Exception as e:
                errors.append(_('Fila %d: %s') % (row_idx, str(e)))
                _logger.error("Error procesando fila %d: %s", row_idx, str(e), exc_info=True)
        
        if errors:
            _logger.warning("Errores al procesar CSV: %s", '\n'.join(errors[:10]))
        
        return lines_data

